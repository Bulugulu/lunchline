"""Rebuild ARTW_model.xlsx as a clean V3-only workbook tied to the latest deck.

Operates IN PLACE on data/research/artw/ARTW_model.xlsx:
  - DROP the legacy v2 tabs (Summary, Scenarios, ValueLenses) and the old
    standalone v3_Summary; rebuild the entangled valuation tabs fresh on v3
    assumptions (WACC 12.5 / 8x / corp-add $0.30M / 3-yr cash-tax holiday).
  - KEEP untouched the already-v3 / factual tabs: SegmentP&L, TwoLeg, Graham,
    Peers, LBO_Returns, LBO_Sensitivity.
  - ADD VCP_Waterfall ($14.0M -> $22.9M, deck slide 8).

Every analytical cell is a live formula referencing yellow inputs; the workbook
is set to full-recalc-on-load so Excel recomputes on open. The script also
recomputes the DCF / SOTP arithmetic in Python and asserts it ties to the
deck's stated $14.5M DCF / $14.0M building / $2.95 base FV before saving.
"""
from pathlib import Path
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = Path("data/research/artw/ARTW_model.xlsx")
DELIV = Path("deliverables/Part1-ARTW-model.xlsx")

NAVY = "0A2540"; RED = "C8102E"; GREY = "6B7280"; YELLOW = "FFF7CC"
GREEN = "1B7A4B"; TINT = "FEF7F7"

f_title = Font(name="Calibri", bold=True, color=NAVY, size=14)
f_sub   = Font(name="Calibri", italic=True, color=GREY, size=9)
f_hdr   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
f_sec   = Font(name="Calibri", bold=True, color=NAVY, size=10)
f_bold  = Font(name="Calibri", bold=True, size=10)
f_red   = Font(name="Calibri", bold=True, color=RED, size=10)
f_grn   = Font(name="Calibri", bold=True, color=GREEN, size=10)
f_note  = Font(name="Calibri", italic=True, color=GREY, size=8)
fill_hdr = PatternFill("solid", fgColor=NAVY)
fill_inp = PatternFill("solid", fgColor=YELLOW)
fill_tint = PatternFill("solid", fgColor=TINT)
thin = Side(style="thin", color="D0D5DC")
border = Border(bottom=thin)
RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left")
WRAP = Alignment(wrap_text=True, vertical="top")

MONEY = '#,##0.00'; ONE = '#,##0.0'; PCT = '0.0%'; PCT0 = '0%'; X = '0.00"x"'


def cell(ws, coord, val, font=None, fill=None, fmt=None, align=None, wrap=False):
    c = ws[coord]
    c.value = val
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = align
    if wrap: c.alignment = WRAP
    return c


def inp(ws, coord, val, fmt=ONE):
    return cell(ws, coord, val, fill=fill_inp, fmt=fmt, align=RIGHT)


# ----------------------------------------------------------------------------
def build_modular_dcf(ws):
    """Standalone modular DCF, 3 scenarios, v3 drivers. Modular EV at row 49."""
    cell(ws, "A1", "Standalone Modular (lab-builder) DCF — v3, 3 scenarios ($M, explicit yr 1–7)", f_title)
    cell(ws, "A2", "v3: WACC 14.0/12.5/11.5; EBITDA margin lifted to 16.6% base (corp-cost add-back $0.45M→$0.30M); "
                   "near-term cash-tax holiday yrs 1–3 (NOL/§382, base & bull). D&A & capex 2.5% rev (asset-light); ΔWC 10% of incremental rev.", f_note)
    cell(ws, "A3", "Driver / Year", f_sec)
    for col, lab in zip("BCD", ("Bear", "Base", "Bull")):
        cell(ws, f"{col}3", lab, f_hdr, fill_hdr, align=RIGHT)

    # inputs
    cell(ws, "A4", "Base-year revenue (FY25 modular)", align=LEFT)
    for col in "BCD": inp(ws, f"{col}4", 10.226)
    growth = {
        "B": [-0.10, -0.05, 0, 0.02, 0.02, 0.02, 0.02],
        "C": [0.08, 0.06, 0.05, 0.04, 0.04, 0.035, 0.03],
        "D": [0.18, 0.14, 0.10, 0.08, 0.06, 0.05, 0.04],
    }
    for i in range(7):
        r = 5 + i
        cell(ws, f"A{r}", f"Yr{i+1} revenue growth", align=LEFT)
        for col in "BCD": inp(ws, f"{col}{r}", growth[col][i], PCT)
    cell(ws, "A12", "EBITDA margin", align=LEFT)
    for col, v in zip("BCD", (0.13, 0.166, 0.19)): inp(ws, f"{col}12", v, PCT)
    cell(ws, "A13", "D&A (% revenue)", align=LEFT)
    for col in "BCD": inp(ws, f"{col}13", 0.025, PCT)
    cell(ws, "A14", "Capex (% revenue)", align=LEFT)
    for col in "BCD": inp(ws, f"{col}14", 0.025, PCT)
    cell(ws, "A15", "ΔWC (% incremental rev)", align=LEFT)
    for col in "BCD": inp(ws, f"{col}15", 0.10, PCT)
    cell(ws, "A16", "WACC", align=LEFT)
    for col, v in zip("BCD", (0.140, 0.125, 0.115)): inp(ws, f"{col}16", v, PCT)
    cell(ws, "A17", "Terminal growth", align=LEFT)
    for col, v in zip("BCD", (0.010, 0.025, 0.030)): inp(ws, f"{col}17", v, PCT)
    cell(ws, "A18", "Cash tax — yrs 1–3", align=LEFT)
    for col, v in zip("BCD", (0.15, 0.0, 0.0)): inp(ws, f"{col}18", v, PCT)
    cell(ws, "A19", "Cash tax — yrs 4–7", align=LEFT)
    for col in "BCD": inp(ws, f"{col}19", 0.15, PCT)

    # projection
    cell(ws, "A21", "PROJECTION ($M)", f_sec)
    # revenue rows 22-28
    for i in range(7):
        r = 22 + i
        cell(ws, f"A{r}", f"Revenue Yr{i+1}", align=LEFT)
        for col in "BCD":
            prev = f"{col}4" if i == 0 else f"{col}{r-1}"
            cell(ws, f"{col}{r}", f"={prev}*(1+{col}{5+i})", fmt=ONE, align=RIGHT)
    # FCF rows 30-36 ; tax = yrs1-3 cell for i<3 else yrs4-7
    cell(ws, "A29", "Free cash flow ($M)", f_sec)
    for i in range(7):
        r = 30 + i
        revr = 22 + i
        taxr = 18 if i < 3 else 19
        cell(ws, f"A{r}", f"FCF Yr{i+1}", align=LEFT)
        for col in "BCD":
            prevrev = f"{col}4" if i == 0 else f"{col}{revr-1}"
            R = f"{col}{revr}"
            m = f"{col}$12"; da = f"{col}$13"; cx = f"{col}$14"; wc = f"{col}$15"; tax = f"{col}${taxr}"
            # FCF = R*m - tax*(R*m - R*da) - R*cx - wc*(R-prevR)
            f = f"={R}*{m}-{tax}*({R}*{m}-{R}*{da})-{R}*{cx}-{wc}*({R}-{prevrev})"
            cell(ws, f"{col}{r}", f, fmt=ONE, align=RIGHT)
    # PV rows 38-44
    cell(ws, "A37", "PV of FCF ($M)", f_sec)
    for i in range(7):
        r = 38 + i; fcfr = 30 + i
        cell(ws, f"A{r}", f"PV FCF Yr{i+1}", align=LEFT)
        for col in "BCD":
            cell(ws, f"{col}{r}", f"={col}{fcfr}/(1+{col}$16)^{i+1}", fmt=ONE, align=RIGHT)
    cell(ws, "A46", "PV explicit FCF (sum)", f_bold)
    for col in "BCD": cell(ws, f"{col}46", f"=SUM({col}38:{col}44)", font=f_bold, fmt=ONE, align=RIGHT)
    cell(ws, "A47", "Terminal value (on Yr7 FCF)", align=LEFT)
    for col in "BCD": cell(ws, f"{col}47", f"={col}36*(1+{col}$17)/({col}$16-{col}$17)", fmt=ONE, align=RIGHT)
    cell(ws, "A48", "PV of terminal value", align=LEFT)
    for col in "BCD": cell(ws, f"{col}48", f"={col}47/(1+{col}$16)^7", fmt=ONE, align=RIGHT)
    cell(ws, "A49", "MODULAR ENTERPRISE VALUE ($M)", f_red)
    for col in "BCD":
        c = cell(ws, f"{col}49", f"={col}46+{col}48", font=f_red, fmt=ONE, align=RIGHT)
        c.fill = fill_tint
    cell(ws, "A51", "Targets (model.md §v3b): DCF EV bear $6.2M / base $14.5M / bull $24.8M.", f_note)
    ws.column_dimensions["A"].width = 30
    for col in "BCD": ws.column_dimensions[col].width = 11


def build_sotp(ws):
    cell(ws, "A1", "Sum-of-the-Parts — carve-out equity (control basis, v3)", f_title)
    cell(ws, "A2", "Building EV = blend of DCF and multiple methods. Carve-out equity = Building EV + net Ag disposal − net debt. $ in millions.", f_note)
    # 1a multiple
    cell(ws, "A4", "1a. BUILDING EV — multiple method", f_sec)
    for col, lab in zip("BCD", ("Bear", "Base", "Bull")): cell(ws, f"{col}5", lab, f_hdr, fill_hdr, align=RIGHT)
    cell(ws, "A6", "Standalone EBITDA used", align=LEFT)
    for col, v in zip("BCD", (1.40, 1.70, 2.00)): inp(ws, f"{col}6", v, MONEY)
    cell(ws, "A7", "EV / EBITDA multiple", align=LEFT)
    for col, v in zip("BCD", (5, 8, 10)): inp(ws, f"{col}7", v, '0.0"x"')
    cell(ws, "A8", "Building EV (multiple)", f_bold)
    for col in "BCD": cell(ws, f"{col}8", f"={col}6*{col}7", font=f_bold, fmt=ONE, align=RIGHT)
    cell(ws, "F6", "EBITDA = seg EBITDA $2.0M − $0.30M standalone corp (v3). Multiple: 8× base "
                   "(Limbach project-era ~9.5×, constr. median ~10×; project-economics cap < that). 5/8/10×.", f_note)
    # 1b DCF leg
    cell(ws, "A10", "1b. BUILDING EV — DCF method (from ModularDCF tab)", f_sec)
    cell(ws, "A11", "Building EV (DCF)", align=LEFT)
    for col in "BCD": cell(ws, f"{col}11", f"=ModularDCF!{col}49", fmt=ONE, align=RIGHT)
    # 1c blend
    cell(ws, "A13", "1c. BUILDING EV — blend (deck value)", f_sec)
    cell(ws, "A14", "Building EV = AVG(multiple, DCF)", f_red)
    for col in "BCD":
        c = cell(ws, f"{col}14", f"=AVERAGE({col}8,{col}11)", font=f_red, fmt=ONE, align=RIGHT); c.fill = fill_tint
    cell(ws, "F14", "Base 8× $13.6M & DCF $14.5M → $14.0M (deck slide 10).", f_note)
    # 1d ag disposal  (Bear% C, Base% D, Bull% E)
    cell(ws, "A16", "1d. AG DISPOSAL — orderly wind-down (recovery haircuts)", f_sec)
    cell(ws, "B17", "Asset $M", f_hdr, fill_hdr, align=RIGHT)
    for col, lab in zip("CDE", ("Bear %", "Base %", "Bull %")): cell(ws, f"{col}17", lab, f_hdr, fill_hdr, align=RIGHT)
    rows = [("Inventory (aged whole-goods)", 9.5, 0.5, 0.6, 0.75),
            ("Accounts receivable", 1.6, 0.85, 0.95, 1.0),
            ("PP&E (single-purpose plant)", 4.5, 0.3, 0.4, 0.55)]
    for i, (lab, asset, b, ba, bu) in enumerate(rows):
        r = 18 + i
        cell(ws, f"A{r}", lab, align=LEFT)
        inp(ws, f"B{r}", asset, ONE)
        for col, v in zip("CDE", (b, ba, bu)): inp(ws, f"{col}{r}", v, PCT0)
    cell(ws, "A21", "Gross disposal proceeds", align=LEFT)
    for col in "CDE": cell(ws, f"{col}21", f"=SUMPRODUCT($B18:$B20,{col}18:{col}20)", fmt=ONE, align=RIGHT)
    cell(ws, "A22", "less: wind-down cost", align=LEFT)
    for col, v in zip("CDE", (-2.0, -1.3, -0.8)): inp(ws, f"{col}22", v, ONE)
    cell(ws, "A23", "Net Ag disposal proceeds", f_bold)
    for col in "CDE": cell(ws, f"{col}23", f"={col}21+{col}22", font=f_bold, fmt=ONE, align=RIGHT)
    cell(ws, "F23", "Base $7.7M (reviewer-disciplined; recoveries 60/95/40%).", f_note)
    # 1e roll-up
    cell(ws, "A25", "1e. SOTP ROLL-UP → carve-out equity / share", f_sec)
    for col, lab in zip("BCD", ("Bear", "Base", "Bull")): cell(ws, f"{col}26", lab, f_hdr, fill_hdr, align=RIGHT)
    cell(ws, "A27", "Building EV (from 1c)", align=LEFT)
    for col in "BCD": cell(ws, f"{col}27", f"={col}14", fmt=ONE, align=RIGHT)
    cell(ws, "A28", "+ Net Ag disposal (from 1d)", align=LEFT)
    for col, src in zip("BCD", "CDE"): cell(ws, f"{col}28", f"={src}23", fmt=ONE, align=RIGHT)
    cell(ws, "A29", "− Net debt at close", align=LEFT)
    for col in "BCD": cell(ws, f"{col}29", "=Summary!$B$8", fmt=ONE, align=RIGHT)
    cell(ws, "A30", "Carve-out equity ($M)", f_bold)
    for col in "BCD": cell(ws, f"{col}30", f"={col}27+{col}28-{col}29", font=f_bold, fmt=ONE, align=RIGHT)
    cell(ws, "A31", "Shares (M, basic)", align=LEFT)
    for col in "BCD": cell(ws, f"{col}31", "=Summary!$B$6", fmt=ONE, align=RIGHT)
    cell(ws, "A32", "FAIR VALUE / SHARE ($)", f_red)
    for col in "BCD":
        c = cell(ws, f"{col}32", f"={col}30/{col}31", font=f_red, fmt=MONEY, align=RIGHT); c.fill = fill_tint
    cell(ws, "A33", "Base ties to deck: $14.0M + $7.7M − $6.4M = $15.3M ÷ 5.184M ≈ $2.95.", f_note)
    ws.column_dimensions["A"].width = 30
    for col in "BCDE": ws.column_dimensions[col].width = 11
    ws.column_dimensions["F"].width = 38


def build_summary(ws):
    cell(ws, "A1", "Art's-Way Manufacturing (ARTW) — V3 Control-Basis Model", f_title)
    cell(ws, "A2", "Single control-basis summary, tied to the Part-1 pitch deck. $ in millions unless per-share. "
                   "Yellow = editable input; everything else is a live formula (recalculates on open).", f_sub)
    # key facts
    cell(ws, "A4", "KEY FACTS  (CITED — FY25 10-K / Q1 FY26 10-Q / 2026 DEF 14A)", f_sec)
    facts = [
        ("Current price ($/sh)", 2.58, MONEY, True),
        ("Shares outstanding (M, basic)", 5.184, ONE, True),
        ("Shares fully diluted (M)", 5.684, ONE, True),
        ("Net debt ($M)", 6.40, ONE, True),
    ]
    r = 5
    for lab, v, fmt, isinp in facts:
        cell(ws, f"A{r}", lab, align=LEFT)
        (inp(ws, f"B{r}", v, fmt) if isinp else cell(ws, f"B{r}", v, fmt=fmt, align=RIGHT))
        r += 1
    cell(ws, "A9", "Market cap ($M)", align=LEFT)
    cell(ws, "B9", "=B5*B6", fmt=ONE, align=RIGHT)
    cell(ws, "A10", "Enterprise value ($M)", align=LEFT)
    cell(ws, "B10", "=B9+B8", fmt=ONE, align=RIGHT)
    cell(ws, "A11", "Book value / share ($)", align=LEFT); inp(ws, "B11", 2.57, MONEY)
    cell(ws, "A12", "Family voting control", align=LEFT); inp(ws, "B12", 0.515, PCT0)
    cell(ws, "A13", "Analyst coverage", align=LEFT); inp(ws, "B13", 0, '0')

    # valuation (live from SOTP)
    cell(ws, "A15", "INTRINSIC VALUE — control basis (live from SOTP)", f_sec)
    cell(ws, "A16", "Scenario", f_hdr, fill_hdr, align=LEFT)
    for col, lab in zip("BCD", ("FV / sh ($)", "vs price", "Prob")): cell(ws, f"{col}16", lab, f_hdr, fill_hdr, align=RIGHT)
    scen = [("Bear", "SOTP!B32", 0.30), ("Base", "SOTP!C32", 0.45), ("Bull", "SOTP!D32", 0.25)]
    for i, (lab, src, p) in enumerate(scen):
        rr = 17 + i
        cell(ws, f"A{rr}", lab, align=LEFT)
        cell(ws, f"B{rr}", f"={src}", fmt=MONEY, align=RIGHT)
        cell(ws, f"C{rr}", f"=B{rr}/$B$5-1", fmt=PCT, align=RIGHT)
        inp(ws, f"D{rr}", p, PCT0)
    cell(ws, "A20", "Probability-weighted", f_bold)
    cell(ws, "B20", "=SUMPRODUCT(B17:B19,D17:D19)", font=f_red, fmt=MONEY, align=RIGHT)
    cell(ws, "C20", "=B20/$B$5-1", font=f_red, fmt=PCT, align=RIGHT)
    cell(ws, "E17", "Base ≈ $2.95–2.96 basic (deck rounds components to $2.95); fully-diluted ≈ $2.70. "
                    "Asset floor: book $2.57, NCAV $1.08 (Graham tab).", f_note)

    # sotp bridge
    cell(ws, "A22", "SOTP BRIDGE — base ($M)", f_sec)
    cell(ws, "A23", "Building business (DCF / multiple blend)", align=LEFT); cell(ws, "B23", "=SOTP!C14", fmt=ONE, align=RIGHT)
    cell(ws, "A24", "+ Ag assets, after orderly sale", align=LEFT); cell(ws, "B24", "=SOTP!D23", fmt=ONE, align=RIGHT)
    cell(ws, "A25", "− Net debt", align=LEFT); cell(ws, "B25", "=-B8", fmt=ONE, align=RIGHT)
    cell(ws, "A26", "Equity ($M)  →  per share", f_bold)
    cell(ws, "B26", "=B23+B24+B25", font=f_bold, fmt=ONE, align=RIGHT)
    cell(ws, "C26", "=B26/B6", font=f_red, fmt=MONEY, align=RIGHT)

    # LBO returns (from LBO_Sensitivity recommended ~$10M row)
    cell(ws, "A28", "SEARCH-FUND LBO RETURNS  (recommended ~$10M debt / 46% lev; 5-yr; from LBO_Sensitivity)", f_sec)
    cell(ws, "A29", "Scenario", f_hdr, fill_hdr, align=LEFT)
    for col, lab in zip("BC", ("MOIC", "IRR")): cell(ws, f"{col}29", lab, f_hdr, fill_hdr, align=RIGHT)
    lbo = [("Bear", "C7", "D7"), ("Base", "E7", "F7"), ("Bull", "G7", "H7")]
    for i, (lab, mo, ir) in enumerate(lbo):
        rr = 30 + i
        cell(ws, f"A{rr}", lab, align=LEFT)
        cell(ws, f"B{rr}", f"=LBO_Sensitivity!{mo}", align=RIGHT)
        cell(ws, f"C{rr}", f"=LBO_Sensitivity!{ir}", align=RIGHT)
    cell(ws, "E30", "Sponsor equity ≈ $11.6M. Deck slide 9 leads with the bull (≈3.7× / ~30% → ~$43M).", f_note)

    # verdict
    cell(ws, "A34", "VERDICT", f_sec)
    cell(ws, "A35",
         "Asymmetric, asset-protected search acquisition. Buy near book on a control basis at ~$2.95 (+14%); the asset "
         "floor (book $2.57, NCAV $1.08, both Graham tests pass) caps the downside, and the return is built by the operator "
         "(divest farm equipment, repay debt, grow the 17%-margin lab builder). Bear is a bounded, collateral-covered loss; "
         "cap leverage near $10M (>55% wipes the bear). Conviction hinges on the undisclosed modular leg-split (data floor).",
         f_note)
    ws["A35"].alignment = WRAP
    ws.merge_cells("A35:H38")
    ws.column_dimensions["A"].width = 34
    for col in "BCD": ws.column_dimensions[col].width = 12
    for col in "EFGH": ws.column_dimensions[col].width = 11


def build_vcp(ws):
    cell(ws, "A1", "Value-Creation Waterfall — building business ($M, control basis)", f_title)
    cell(ws, "A2", "Deck slide 8. Three operator levers the buyer actually builds, plus a smaller market re-rating. "
                   "Driver assumptions benchmarked in model.md §v4; the high-/low-end leg split is undisclosed (mix lever is an estimate).", f_note)
    cell(ws, "A4", "Lever", f_hdr, fill_hdr, align=LEFT)
    cell(ws, "B4", "Value add ($M)", f_hdr, fill_hdr, align=RIGHT)
    cell(ws, "C4", "Bucket", f_hdr, fill_hdr, align=LEFT)
    cell(ws, "D4", "Driver", f_hdr, fill_hdr, align=LEFT)
    cell(ws, "A5", "Building business today (control basis)", f_bold)
    cell(ws, "B5", "=SOTP!C14", font=f_bold, fmt=ONE, align=RIGHT)
    rows = [
        ("Grow sales", 3.3, "Commercial", "Real head of sales lifts growth ~2–3% → ~6–8%/yr (backlog +103% under a part-time president)."),
        ("Move up-market", 2.4, "Mix shift", "Shift to BSL-3/GMP (2–3× $/sq ft) lifts blended op margin ~17% → ~19–21%."),
        ("Upsell the installed base", 1.7, "Service tail", "Annual re-certification/monitoring resold to 150+ delivered buildings (~$7B fragmented services)."),
        ("Re-rate (a market call)", 1.5, "Re-rate", "Standalone faster-growing biocontainment builder toward ~10× sector median. Sized smallest."),
    ]
    r = 6
    for lab, v, bucket, driver in rows:
        cell(ws, f"A{r}", lab, align=LEFT)
        inp(ws, f"B{r}", v, ONE)
        cell(ws, f"C{r}", bucket, align=LEFT)
        cell(ws, f"D{r}", driver, align=LEFT, wrap=True)
        r += 1
    cell(ws, f"A{r}", "Building business if all four work", f_grn)
    cell(ws, f"B{r}", f"=B5+SUM(B6:B{r-1})", font=f_grn, fmt=ONE, align=RIGHT)
    ws[f"B{r}"].fill = fill_tint
    cell(ws, f"A{r+2}", "Target ties to deck: $14.0M → $22.9M.", f_note)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 70


def build_assumptions(ws):
    cell(ws, "A1", "Assumptions & judgment log — v3 control basis", f_title)
    cell(ws, "A2", "How each non-cited input was built. CITED = from filings; ESTIMATED/JUDGMENT = reasoned. Full derivation in "
                   "model.md §v3a, wacc_research.md, filing_review.md, peer_notes.md.", f_note)
    cell(ws, "A4", "Input", f_hdr, fill_hdr, align=LEFT)
    cell(ws, "B4", "v3 base", f_hdr, fill_hdr, align=RIGHT)
    cell(ws, "C4", "Tier", f_hdr, fill_hdr, align=LEFT)
    cell(ws, "D4", "How constructed / logic", f_hdr, fill_hdr, align=LEFT)
    rows = [
        ("WACC", "12.5%", "Built (CAPM + premia)", "rf 4.3% + β1.4×ERP 5.0% + ~3% size/illiquidity. Control buyer drops the marketable-minority illiquidity slice (v2 13.5%→12.5%). Bear 14.0% / bull 11.5%."),
        ("Modular EV/EBITDA multiple", "8×", "Judgment (comps)", "Limbach project-era ~9.5×, construction median ~10×; project economics (no recurring/leasing) cap below that. Range 6–9×. Bear 5× / bull 10×."),
        ("Standalone modular EBITDA", "$1.70M", "Built", "Segment EBITDA $2.0M − $0.30M standalone corp add-back (take-private removes public-co cost; v2 $0.45M→$0.30M). → DCF margin ~16.6%."),
        ("Near-term cash tax", "~0% yrs 1–3", "CITED + judgment", "~$7.1M gross federal NOL, no valuation allowance (filing_review.md); §382 caps usage post-control, so only a modest 3-yr holiday claimed, then 15%. Bear claims none."),
        ("Ag inventory recovery", "60%", "ESTIMATED", "Aged whole-goods into a slow market; beet-equipment hit by Dec-2025 ACS payment cut. 55–65% realistic. Bear 50% / bull 75%."),
        ("Ag AR recovery", "95%", "ESTIMATED", "Trade AR collects near par in an orderly wind-down. Bear 85% / bull 100%."),
        ("Ag PP&E recovery", "40%", "ESTIMATED", "Single-purpose rural-Iowa plant; 2024 Tools real-estate sale a loose datapoint. Bear 30% / bull 55%."),
        ("Ag wind-down cost", "$1.3M", "ESTIMATED", "Severance, lease breakage, warranty/dealer tail, transaction. Bear $2.0M / bull $0.8M."),
        ("D&A & capex (% rev)", "2.5% each", "Judgment", "Asset-light modular ($3.3M assets, $0.2M capex on $10.2M rev)."),
        ("ΔWC (% incremental rev)", "10%", "Judgment", "Project-timing-driven; nets ~zero over a cycle, not a permanent build."),
        ("Terminal growth", "2.5%", "Judgment", "GDP-ish. Bear 1.0% / bull 3.0%."),
        ("Scenario probabilities", "30/45/25", "Judgment", "Bear-weighted for the data-floor-limited, leg-split-dependent setup."),
        ("Two-leg split (research/ag-bio)", "60/40", "ESTIMATED — DATA FLOOR", "From the single '+$1.355M ag-building' breadcrumb. THE key undisclosed assumption (see TwoLeg tab)."),
        ("Net debt", "$6.40M", "CITED", "Total debt $6.41M (revolver + term + finance leases incl SBA EIDL) − cash $0.005M."),
        ("LBO entry premium", "~10%", "CITED precedent", "Friendly control freeze-out comps (not the v2 31.5% FONR template, which is superseded). Detail in LBO_Returns."),
    ]
    r = 5
    for lab, v, tier, logic in rows:
        cell(ws, f"A{r}", lab, align=LEFT)
        cell(ws, f"B{r}", v, align=RIGHT)
        cell(ws, f"C{r}", tier, align=LEFT)
        cell(ws, f"D{r}", logic, align=LEFT, wrap=True)
        ws.row_dimensions[r].height = 26
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 78


# ----------------------------------------------------------------------------
def verify():
    """Independently recompute the DCF/SOTP arithmetic and assert it ties."""
    def dcf(rev0, g, m, wacc, tg, tax, da=0.025, cx=0.025, dwc=0.10):
        revs = []; prev = rev0
        for gi in g: prev *= (1 + gi); revs.append(prev)
        fcf = []; prev = rev0
        for i, R in enumerate(revs):
            ti = tax[i]
            fcf.append(R*m - ti*(R*m - R*da) - R*cx - dwc*(R-prev)); prev = R
        pv = sum(f/(1+wacc)**(i+1) for i, f in enumerate(fcf))
        tv = fcf[-1]*(1+tg)/(wacc-tg)
        return pv + tv/(1+wacc)**7
    hol = [0,0,0,0.15,0.15,0.15,0.15]; flat = [0.15]*7
    base_dcf = dcf(10.226, [0.08,0.06,0.05,0.04,0.04,0.035,0.03], 0.166, 0.125, 0.025, hol)
    base_mult = 8*1.70
    base_bldg = (base_dcf + base_mult)/2
    base_ag = 0.6*9.5 + 0.95*1.6 + 0.4*4.5 - 1.3
    base_eq = base_bldg + base_ag - 6.40
    base_fv = base_eq/5.184
    print(f"  base DCF EV   = {base_dcf:5.2f}  (deck 14.5)")
    print(f"  base mult EV  = {base_mult:5.2f}  (deck 13.6)")
    print(f"  base building = {base_bldg:5.2f}  (deck 14.0)")
    print(f"  base ag       = {base_ag:5.2f}  (deck 7.7)")
    print(f"  base equity   = {base_eq:5.2f}  (deck 15.3)")
    print(f"  base FV/sh    = {base_fv:5.2f}  (deck 2.95)")
    assert abs(base_dcf-14.5) < 0.1, base_dcf
    assert abs(base_mult-13.6) < 0.01
    assert abs(base_ag-7.7) < 0.05
    assert abs(base_fv-2.95) < 0.03, base_fv
    print("  ✓ ties to deck within rounding")


def main():
    verify()
    wb = openpyxl.load_workbook(SRC)
    # drop legacy + tabs we rebuild
    # idempotent: drop legacy AND any tab we rebuild (incl. on a re-run)
    for name in ("v3_Summary", "Summary", "Scenarios", "ValueLenses", "SOTP",
                 "ModularDCF", "Assumptions", "VCP_Waterfall"):
        if name in wb.sheetnames:
            del wb[name]
    # (Assumptions intentionally dropped — its v3 build now lives inline in ModularDCF/SOTP/Summary notes)
    ws_sum = wb.create_sheet("Summary"); build_summary(ws_sum)
    ws_sotp = wb.create_sheet("SOTP"); build_sotp(ws_sotp)
    ws_dcf = wb.create_sheet("ModularDCF"); build_modular_dcf(ws_dcf)
    ws_vcp = wb.create_sheet("VCP_Waterfall"); build_vcp(ws_vcp)
    ws_asm = wb.create_sheet("Assumptions"); build_assumptions(ws_asm)
    # order
    order = ["Summary", "SegmentP&L", "SOTP", "ModularDCF", "VCP_Waterfall",
             "TwoLeg", "Graham", "Peers", "Assumptions", "LBO_Returns", "LBO_Sensitivity"]
    order = [n for n in order if n in wb.sheetnames] + [n for n in wb.sheetnames if n not in order]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    wb.active = 0
    # recalc on open
    wb.calculation.fullCalcOnLoad = True
    wb.save(SRC)
    shutil.copyfile(SRC, DELIV)
    print(f"\nSaved {SRC} and copied to {DELIV}")
    print("Tabs:", wb.sheetnames)


if __name__ == "__main__":
    main()
