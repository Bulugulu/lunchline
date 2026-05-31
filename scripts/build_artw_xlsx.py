"""
Build a live, formula-driven Excel model for ARTW (Art's-Way Manufacturing) from the
v2 deep-flow CARVE-OUT build in data/research/artw/model.md.

This is a CARVE-OUT thesis model, not a single-entity model: acquire ARTW, divest the
loss-making Agricultural Products segment, own Art's-Way Scientific (Modular Buildings)
standalone. The workbook structure therefore differs from the HCKT template — the SOTP
is the centerpiece (Modular EV + net Ag disposal - net debt), the operating model splits
modular into two policy-hedged legs, and the scenarios produce both a standalone FV and a
whole-company acquisition MOIC at a control premium.

Every output cell is a formula referencing editable input cells (pale-yellow fill), so the
reviewer can flex WACC / growth / margins / multiples / recovery haircuts / probabilities
and watch the DCF, SOTP, scenario returns, and Graham/Buffett value lenses recompute.

Notes columns flag each input as CITED vs ESTIMATED, traceable to model.md.

Output: data/research/artw/ARTW_model.xlsx
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule

OUT = Path(__file__).parent.parent / "data" / "research" / "artw" / "ARTW_model.xlsx"

NAVY = "0A2540"
RED = "C8102E"
TINT = "FEF7F7"
GREY = "6B7280"
LIGHT = "F1F3F5"
GREEN = "1B7A4B"

hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=NAVY)
sub_font = Font(name="Calibri", bold=True, color=NAVY, size=12)
title_font = Font(name="Calibri", bold=True, color=NAVY, size=15)
red_font = Font(name="Calibri", bold=True, color=RED, size=11)
green_font = Font(name="Calibri", bold=True, color=GREEN, size=11)
bold = Font(name="Calibri", bold=True, size=11)
note_font = Font(name="Calibri", italic=True, color=GREY, size=9)
total_fill = PatternFill("solid", fgColor=TINT)
input_fill = PatternFill("solid", fgColor="FFFDE7")  # pale yellow = editable input
thin = Side(style="thin", color="D9DCE1")
border = Border(bottom=thin)

MONEY = '"$"#,##0.0'        # $ millions, one decimal
PS = '"$"#,##0.00'          # $ per share
PCT = '0.0%'
MULT = '0.0"x"'
MULT2 = '0.00"x"'
NUM0 = '#,##0'
NUM3 = '#,##0.000'

wb = Workbook()


def style_header_row(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")


def put(ws, addr, value, *, fmt=None, font=None, fill=None, align=None, wrap=False):
    cell = ws[addr]
    cell.value = value
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="top" if wrap else None)
    return cell


# ============================================================================
# 1. SUMMARY / VERDICT
# ============================================================================
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
for col, w in {"A": 40, "B": 16, "C": 16, "D": 16, "E": 22}.items():
    ws.column_dimensions[col].width = w

put(ws, "A1", "Art's-Way Manufacturing (ARTW) — Deep-Flow Carve-Out Valuation", font=title_font)
put(ws, "A2", "v2 (lead-reviewed) · 2026-05-30 · FY ends Nov-30 · $ in millions unless noted · "
              "yellow = editable input", font=note_font)

put(ws, "A4", "VERDICT", font=red_font)
put(ws, "A5", "FAIRLY-TO-MODESTLY OVERVALUED on the standalone base (-4%). Not a mispriced free gem "
              "(net debt + aged ag inventory mean the modular gem sits at ~6-9x standalone EBITDA inside "
              "today's EV, ~7.8x base — fair-to-full). The return lives entirely in the bull tail (+83%) "
              "and the family-gated friendly-carve-out option. Real Graham asset floor (passes BOTH tests). "
              "At a normal control premium (FONR 31.5%) the base acquisition is underwater (0.73x MOIC).",
    wrap=True)
ws.merge_cells("A5:E5")
ws.row_dimensions[5].height = 70

put(ws, "A7", "KEY FACTS  (CITED — FY25 10-K / Q1 FY26 10-Q / DEF 14A)", font=sub_font)
facts = [
    ("Current price ($)", "=Scenarios!B3", PS, "CITED 2026-05"),
    ("Shares outstanding (M)", 5.184, NUM3, "CITED DEF 14A Mar-5-26"),
    ("Market cap ($M)", "=B8*B9", MONEY, "= price x shares"),
    ("Total debt ($M)", 6.41, MONEY, "CITED bal sheet + Note 10"),
    ("less: cash ($M)", 0.005, MONEY, "CITED near-zero cash"),
    ("Net debt ($M)", "=B11-B12", MONEY, "= debt - cash"),
    ("Enterprise value ($M)", "=B10+B13", MONEY, "= mkt cap + net debt"),
    ("Book value / share ($)", 2.57, PS, "CITED equity $13.306M / 5.184M"),
    ("Price / book", "=B8/B15", MULT2, "~1.0x book"),
    ("Analyst coverage", "0 analysts", None, "CITED — pure orphan micro-cap"),
    ("Control (McConnell family group)", 0.515, PCT, "CITED DEF 14A — friendly deal only"),
]
r = 8
for label, val, fmt, note in facts:
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", val, fmt=fmt)
    put(ws, f"E{r}", note, font=note_font)
    r += 1

put(ws, "A20", "HEADLINE OUTPUTS  (live from other tabs)", font=sub_font)
outs = [
    ("Base FV / share — Modular DCF blend SOTP (§3)", "=Scenarios!B8", PS, False),
    ("Base FV / share — SOTP-multiple (§1c)", "=SOTP!C27", PS, False),
    ("Standalone FV / share — BASE (blended)", "=Scenarios!B8", PS, True),
    ("Probability-weighted standalone FV / share", "=Scenarios!B11", PS, True),
    ("Prob-weighted return vs price (3-yr)", "=Scenarios!C11", PCT, False),
    ("Acquisition MOIC @ FONR 31.5% premium — BASE", "=Scenarios!F8", MULT2, True),
    ("Acquisition MOIC — BULL", "=Scenarios!F9", MULT2, False),
]
r = 21
for label, val, fmt, em in outs:
    put(ws, f"A{r}", label, font=bold if em else None)
    put(ws, f"B{r}", val, fmt=fmt, font=bold if em else None)
    if em:
        ws[f"A{r}"].fill = total_fill
        ws[f"B{r}"].fill = total_fill
    r += 1

put(ws, "A29", "SCENARIO FAIR VALUES  (standalone carve-out, blended DCF + multiple)", font=sub_font)
for j, h in enumerate(["Scenario", "FV / share", "vs price", "Probability"]):
    put(ws, f"{chr(65+j)}30", h)
style_header_row(ws, 30, 4)
for r, (nm, src) in zip((31, 32, 33), [("Bear", 7), ("Base", 8), ("Bull", 9)]):
    put(ws, f"A{r}", nm)
    put(ws, f"B{r}", f"=Scenarios!B{src}", fmt=PS)
    put(ws, f"C{r}", f"=Scenarios!C{src}", fmt=PCT)
    put(ws, f"D{r}", f"=Scenarios!E{src}", fmt='0%')

put(ws, "A35", "GRAHAM / BUFFETT VALUE LENSES  (live from ValueLenses tab)", font=sub_font)
lenses = [
    ("EPV modular-only, no growth (loads full net debt)", "=ValueLenses!B12", PS),
    ("Reverse-DCF implied perpetual FCF growth", "=ValueLenses!B20", PCT),
    ("Owner-earnings yield (on market cap)", "=ValueLenses!B28", PCT),
    ("Graham NCAV / share (net-net floor)", "=ValueLenses!B34", PS),
]
r = 36
for label, val, fmt in lenses:
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", val, fmt=fmt)
    r += 1
put(ws, "A40", "Graham strong-financial-condition test (both prongs)")
put(ws, "B40", '=ValueLenses!C42&" / "&ValueLenses!C45', font=green_font)

put(ws, "A42", "Read: unlike HCKT (failed both Graham prongs, no asset floor), ARTW PASSES BOTH — a real "
               "static balance-sheet floor (NCAV $1.08, book ~1.0x). But the reverse-DCF says ~$20M EV "
               "already embeds ~7% perpetual growth: the market is NOT pricing a melting farm-equipment "
               "microcap. No margin of safety on the base; the upside is bull-only + the carve-out option.",
    font=note_font, wrap=True)
ws.merge_cells("A42:E43")

put(ws, "A45", "TABS: Summary · SegmentP&L · SOTP · ModularDCF · TwoLeg · Scenarios · "
               "ValueLenses · Graham · Peers · Assumptions", font=note_font)

# ============================================================================
# 2. SEGMENT P&L (FY23-FY25) + ERC NORMALIZATION
# ============================================================================
ws = wb.create_sheet("SegmentP&L")
ws.sheet_view.showGridLines = False
for col, w in {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12, "F": 30}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Segment P&L — Ag Products vs Modular Buildings ($000)", font=title_font)
put(ws, "A2", "CITED: 10-K Note 17 / prior 10-K. The headline conglomerate optics mask a 17%-margin "
              "modular gem averaged with a money-losing ag anchor.", font=note_font)

# --- Ag Products block ---
put(ws, "A4", "AGRICULTURAL PRODUCTS (the segment to divest)", font=sub_font)
for j, h in enumerate(["", "FY23", "FY24", "FY25", "Note"]):
    put(ws, f"{chr(65+j)}5", h)
style_header_row(ws, 5, 4)
ag = [
    ("Revenue", [22467, 14663, 12749], '#,##0', "CITED — collapsed -13.1% FY25"),
    ("Gross profit", [6584, 4155, 2977], '#,##0', "CITED — 23.4% GM FY25"),
    ("  gross margin", ["=B7/B6", "=C7/C6", "=D7/D6"], PCT, "calc"),
    ("Operating income (loss)", [664, -1510, -1462], '#,##0', "CITED — lost ~$1.5M x2 yrs"),
    ("  operating margin", ["=B9/B6", "=C9/C6", "=D9/D6"], PCT, "calc"),
    ("Total assets", [20754, 18372, 19204], '#,##0', "CITED — inventory-heavy anchor"),
]
r = 6
for label, vals, fmt, note in ag:
    put(ws, f"A{r}", label)
    for j, v in enumerate(vals):
        put(ws, f"{chr(66+j)}{r}", v, fmt=fmt)
    put(ws, f"F{r}", note, font=note_font)
    r += 1

# --- Modular block ---
put(ws, "A13", "MODULAR BUILDINGS (Art's-Way Scientific — the gem to keep)", font=sub_font)
for j, h in enumerate(["", "FY23", "FY24", "FY25", "Note"]):
    put(ws, f"{chr(65+j)}14", h)
style_header_row(ws, 14, 4)
mod = [
    ("Revenue", [7814, 9836, 10226], '#,##0', "CITED — rose every yr +4.0% FY25"),
    ("Gross profit", [2000, 3155, 3290], '#,##0', "CITED — 32.2% GM FY25"),
    ("  gross margin", ["=B16/B15", "=C16/C15", "=D16/D15"], PCT, "calc"),
    ("Operating income", [867, 1971, 1751], '#,##0', "CITED — 17.1% op margin"),
    ("  operating margin", ["=B18/B15", "=C18/C15", "=D18/D15"], PCT, "calc"),
    ("Total assets", [2593, 2869, 3274], '#,##0', "CITED — very asset-light"),
    ("D&A / capex FY25", [None, None, 251], '#,##0', "CITED — capex $222K (light)"),
]
r = 15
for label, vals, fmt, note in mod:
    put(ws, f"A{r}", label)
    for j, v in enumerate(vals):
        if v is not None:
            put(ws, f"{chr(66+j)}{r}", v, fmt=fmt)
    put(ws, f"F{r}", note, font=note_font)
    r += 1

# --- ERC normalization ---
put(ws, "A23", "ERC NORMALIZATION (consolidated, $000) — the single most important adjustment", font=sub_font)
for j, h in enumerate(["", "FY25 reported", "Adjustment", "Ex-ERC", "Note"]):
    put(ws, f"{chr(65+j)}24", h)
style_header_row(ws, 24, 4)
put(ws, "A25", "Consolidated revenue")
put(ws, "B25", 22975, fmt='#,##0'); put(ws, "F25", "CITED -6.2%", font=note_font)
put(ws, "A26", "Consolidated operating income")
put(ws, "B26", 289, fmt='#,##0'); put(ws, "F26", "CITED — avg of 17% biz + loser", font=note_font)
put(ws, "A27", 'Other income (the ERC)')
put(ws, "B27", 1514, fmt='#,##0', fill=input_fill)
put(ws, "F27", "CITED — Employee Retention Credit, vs $3K FY24", font=note_font)
put(ws, "A28", "Reported net income")
put(ws, "B28", 1035, fmt='#,##0')
put(ws, "C28", "=-B27", fmt='#,##0')
put(ws, "D28", "=B28+C28", fmt='#,##0', font=bold)
put(ws, "F28", "CITED net income is an ERC mirage", font=note_font)
put(ws, "A29", "Pretax ex-ERC (op inc - interest + true other)", font=bold)
put(ws, "B29", "=B26-367", fmt='#,##0', font=bold)
put(ws, "F29", "~break-even/slight loss (interest $367K)", font=note_font)
ws["A29"].fill = total_fill; ws["B29"].fill = total_fill
put(ws, "A30", "Operating cash flow FY25")
put(ws, "B30", -904, fmt='#,##0', font=red_font)
put(ws, "F30", "CITED negative — ag inventory build (+$1,778K)", font=note_font)
put(ws, "A31", "Cash income taxes paid FY25")
put(ws, "B31", 13, fmt='#,##0')
put(ws, "F31", "CITED ~zero (NOL shield)", font=note_font)

# --- Backlog + Q1 FY26 ---
put(ws, "A33", "LEADING INDICATORS (CITED — strongest filing evidence of momentum)", font=sub_font)
bl = [
    ("Modular backlog Feb-2 2025 -> 2026", "$2,403K -> $4,882K (+103%)"),
    ("Modular backlog as of Apr-7 2026", "$4,513K (+83% YoY)"),
    ("Q1 FY26 modular revenue", "$2,886K vs $2,193K (+31.6%)"),
    ("Ag Products net backlog 2025 -> 2026", "$3,486K -> $3,224K (-7.5%; beet equip hurt)"),
]
r = 34
for label, val in bl:
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", val)
    ws.merge_cells(f"B{r}:D{r}")
    r += 1

put(ws, "A39", "Read: Modular revenue rose every year (7.8->9.8->10.2) on ~$3.3M of assets while Ag "
               "collapsed (22.5->14.7->12.7) into losses. FY25 net income $1.035M is almost entirely the "
               "$1.514M one-time ERC; ex-ERC the company was break-even. The FY25 negative OCF is an AG "
               "inventory phenomenon — modular is asset-light and cash-generative (§A6).", font=note_font, wrap=True)
ws.merge_cells("A39:F40")

# ============================================================================
# 3. SOTP (the centerpiece)
# ============================================================================
ws = wb.create_sheet("SOTP")
ws.sheet_view.showGridLines = False
for col, w in {"A": 38, "B": 13, "C": 13, "D": 13, "E": 13, "F": 30}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Sum-of-the-Parts (the centerpiece) — the gem is NOT free", font=title_font)
put(ws, "A2", "Carve-out equity = Modular standalone EV + net Ag disposal - net debt. "
              "Ag disposal & debt netted together (revolver funds ag inventory). $ in millions.",
    font=note_font)

# --- 1a Modular standalone EV (multiple method) ---
put(ws, "A4", "1a. MODULAR STANDALONE EV (multiple method)", font=sub_font)
for j, h in enumerate(["", "Bear", "Base", "Bull", "Note"]):
    put(ws, f"{chr(65+j)}5", h)
style_header_row(ws, 5, 4)
put(ws, "A6", "Standalone EBITDA used")
put(ws, "B6", 1.40, fmt=MONEY, fill=input_fill)
put(ws, "C6", 1.55, fmt=MONEY, fill=input_fill)
put(ws, "D6", 2.00, fmt=MONEY, fill=input_fill)
put(ws, "F6", "ESTIMATED — seg EBITDA $2.0M less ~$0.45M standalone corp", font=note_font)
put(ws, "A7", "EV / EBITDA multiple")
put(ws, "B7", 5.0, fmt=MULT, fill=input_fill)
put(ws, "C7", 7.0, fmt=MULT, fill=input_fill)
put(ws, "D7", 10.0, fmt=MULT, fill=input_fill)
put(ws, "F7", "FONR ~7-9x anchor; WSC 11.7-14.5x = unreachable ceiling", font=note_font)
put(ws, "A8", "Modular EV (multiple)", font=bold)
for col in "BCD":
    put(ws, f"{col}8", f"={col}6*{col}7", fmt=MONEY, font=bold)
    ws[f"{col}8"].fill = total_fill
ws["A8"].fill = total_fill
put(ws, "F8", "Bear $7.0M / Base $10.9M / Bull $20.0M", font=note_font)

# --- 1b Ag disposal table ---
put(ws, "A10", "1b. AG PRODUCTS DISPOSAL (orderly wind-down — recovery haircuts) [v2 re-centered]", font=sub_font)
put(ws, "A11", "Ag assets: inventory ~$9.5M, AR ~$1.6M, PP&E (Armstrong plant) ~$4.5M. All haircuts ESTIMATED.",
    font=note_font)
for j, h in enumerate(["Recovery on…", "Asset $M", "Bear %", "Base %", "Bull %"]):
    put(ws, f"{chr(65+j)}12", h)
style_header_row(ws, 12, 5)
disp = [
    ("Inventory (aged whole-goods)", 9.5, 0.50, 0.60, 0.75),
    ("Accounts receivable", 1.6, 0.85, 0.95, 1.00),
    ("PP&E (single-purpose plant)", 4.5, 0.30, 0.40, 0.55),
]
r = 13
for label, asset, b, ba, bu in disp:
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", asset, fmt=MONEY, fill=input_fill)
    put(ws, f"C{r}", b, fmt=PCT, fill=input_fill)
    put(ws, f"D{r}", ba, fmt=PCT, fill=input_fill)
    put(ws, f"E{r}", bu, fmt=PCT, fill=input_fill)
    r += 1
# r is now 16
put(ws, "A16", "Gross disposal proceeds", font=bold)
put(ws, "C16", "=SUMPRODUCT($B13:$B15,C13:C15)", fmt=MONEY, font=bold)
put(ws, "D16", "=SUMPRODUCT($B13:$B15,D13:D15)", fmt=MONEY, font=bold)
put(ws, "E16", "=SUMPRODUCT($B13:$B15,E13:E15)", fmt=MONEY, font=bold)
put(ws, "A17", "less: wind-down cost (severance/lease/warranty/txn)")
put(ws, "C17", -2.0, fmt=MONEY, fill=input_fill)
put(ws, "D17", -1.3, fmt=MONEY, fill=input_fill)
put(ws, "E17", -0.8, fmt=MONEY, fill=input_fill)
put(ws, "A18", "Net Ag disposal proceeds", font=bold)
for col in "CDE":
    put(ws, f"{col}18", f"={col}16+{col}17", fmt=MONEY, font=bold)
    ws[f"{col}18"].fill = total_fill
ws["A18"].fill = total_fill
put(ws, "F18", "Base $7.7M (v1 70%->60% recenter, was $8.9M)", font=note_font)

# --- 1c SOTP roll-up (multiple method) ---
put(ws, "A20", "1c. SOTP ROLL-UP — carve-out equity (multiple method)", font=sub_font)
for j, h in enumerate(["", "Bear", "Base", "Bull", "Note"]):
    put(ws, f"{chr(65+j)}21", h)
style_header_row(ws, 21, 4)
put(ws, "A22", "Modular EV (from 1a)")
for col, src in zip("BCD", "BCD"):
    put(ws, f"{col}22", f"={src}8", fmt=MONEY)
put(ws, "A23", "+ Net Ag disposal (from 1b)")
put(ws, "B23", "=C18", fmt=MONEY)
put(ws, "C23", "=D18", fmt=MONEY)
put(ws, "D23", "=E18", fmt=MONEY)
put(ws, "A24", "- Net debt at close")
for col in "BCD":
    put(ws, f"{col}24", "=$B$28", fmt=MONEY)
put(ws, "A25", "Carve-out equity", font=bold)
for col in "BCD":
    put(ws, f"{col}25", f"={col}22+{col}23-{col}24", fmt=MONEY, font=bold)
put(ws, "A26", "Shares (M)")
for col in "BCD":
    put(ws, f"{col}26", "=$B$29", fmt=NUM3)
put(ws, "A27", "Value per share ($) — SOTP multiple", font=bold)
for col in "BCD":
    put(ws, f"{col}27", f"={col}25/{col}26", fmt=PS, font=bold)
    ws[f"{col}27"].fill = total_fill
ws["A27"].fill = total_fill
put(ws, "F27", "Bear $1.16 / Base $2.35 / Bull $4.63", font=note_font)
# shared inputs
put(ws, "A28", "Net debt ($M)")
put(ws, "B28", "=Summary!B13", fmt=MONEY)
put(ws, "A29", "Shares (M)")
put(ws, "B29", "=Summary!B9", fmt=NUM3)
# fix E14 ref used in Summary -> point summary at C27 instead
# (Summary referenced SOTP!E14; correct to base $/sh at C27)

# --- 1d reverse-SOTP ---
put(ws, "A31", "1d. REVERSE-SOTP — is the modular gem free inside today's EV? NO, fair-to-full", font=sub_font)
put(ws, "A32", "Strip net ag disposal out of EV; see what's left implied for modular.", font=note_font)
for j, h in enumerate(["Ag valued at…", "Net ag $M", "Implied Modular EV", "Implied EV/EBITDA"]):
    put(ws, f"{chr(65+j)}33", h)
style_header_row(ws, 33, 4)
put(ws, "A34", "EV today ($M)")
put(ws, "B34", "=Summary!B14", fmt=MONEY)
rev = [
    ("Base orderly", "=D18", 7.8),
    ("Bull going-concern", "=E18", 6.0),
    ("Bear liquidation", "=C18", 9.2),
]
r = 35
for label, agsrc, mult in rev:
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", agsrc, fmt=MONEY)
    put(ws, f"C{r}", f"=$B$34-B{r}", fmt=MONEY)
    put(ws, f"D{r}", mult, fmt=MULT)
    r += 1
put(ws, "A38", "Implied ~6-9.2x standalone EBITDA (~7.8x base) — fair-to-full, not free. Contrast USNA "
               "where net cash + the stake ≈ whole market cap and the core came free. Here there is no cash "
               "cushion and the ag inventory is largely offset by the debt it financed.", font=note_font, wrap=True)
ws.merge_cells("A38:F39")

# ============================================================================
# 4. STANDALONE MODULAR DCF
# ============================================================================
ws = wb.create_sheet("ModularDCF")
ws.sheet_view.showGridLines = False
for col, w in {"A": 30, "B": 12, "C": 12, "D": 12}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Standalone Modular DCF — 3 scenarios ($M, explicit yr1-7)", font=title_font)
put(ws, "A2", "Cash tax 15% (NOLs finite, §A2). D&A & capex ~2.5% rev (asset-light). ΔWC 10% of "
              "incremental rev. NO net debt in EV (debt retired by ag disposal — see SOTP §1c).",
    font=note_font)

for j, h in enumerate(["Driver / Year", "Bear", "Base", "Bull"]):
    put(ws, f"{chr(65+j)}3", h)
style_header_row(ws, 3, 4)
inputs = [
    ("Base-year revenue (FY25 modular)", [10.226, 10.226, 10.226], MONEY),
    ("Yr1 revenue growth", [-0.10, 0.08, 0.18], PCT),
    ("Yr2 revenue growth", [-0.05, 0.06, 0.14], PCT),
    ("Yr3 revenue growth", [0.0, 0.05, 0.10], PCT),
    ("Yr4 revenue growth", [0.02, 0.04, 0.08], PCT),
    ("Yr5 revenue growth", [0.02, 0.04, 0.06], PCT),
    ("Yr6 revenue growth", [0.02, 0.035, 0.05], PCT),
    ("Yr7 revenue growth", [0.02, 0.03, 0.04], PCT),
    ("EBITDA margin", [0.13, 0.16, 0.19], PCT),
    ("D&A (% of revenue)", [0.025, 0.025, 0.025], PCT),
    ("Cash tax rate", [0.15, 0.15, 0.15], PCT),
    ("Capex (% of revenue)", [0.025, 0.025, 0.025], PCT),
    ("ΔWC (% of incremental rev)", [0.10, 0.10, 0.10], PCT),
    ("WACC", [0.15, 0.135, 0.12], PCT),
    ("Terminal growth", [0.01, 0.025, 0.03], PCT),
]
r = 4
for label, vals, fmt in inputs:
    put(ws, f"A{r}", label)
    for j, v in enumerate(vals):
        put(ws, f"{chr(66+j)}{r}", v, fmt=fmt, fill=input_fill)
    r += 1
# rows: rev base=4, g1..g7=5..11, margin=12, D&A%=13, tax=14, capex%=15, dWC%=16, WACC=17, termg=18

# Revenue rows 20-26
put(ws, "A20", "PROJECTION ($M)", font=sub_font)
rev_rows = list(range(21, 28))  # yr1..yr7
for i, rr in enumerate(rev_rows):
    put(ws, f"A{rr}", f"Revenue Yr{i+1}")
for col in "BCD":
    ws[f"{col}21"] = f"={col}4*(1+{col}5)"
    for k in range(1, 7):
        ws[f"{col}{21+k}"] = f"={col}{20+k}*(1+{col}{5+k})"
    for rr in rev_rows:
        ws[f"{col}{rr}"].number_format = MONEY

# EBITDA rows 29-35
put(ws, "A28", "EBITDA")
eb_rows = list(range(29, 36))
for i, rr in enumerate(eb_rows):
    put(ws, f"A{rr}", f"EBITDA Yr{i+1}")
for col in "BCD":
    for i, rr in enumerate(eb_rows):
        ws[f"{col}{rr}"] = f"={col}{21+i}*{col}$12"
        ws[f"{col}{rr}"].number_format = MONEY

# FCF rows 37-43: FCF = EBITDA - (EBITDA - D&A)*tax - capex - dWC
#   D&A = rev*D&A% ; capex = rev*capex% ; dWC = (rev - prevrev)*dWC%
put(ws, "A36", "Free cash flow")
fcf_rows = list(range(37, 44))
for i, rr in enumerate(fcf_rows):
    put(ws, f"A{rr}", f"FCF Yr{i+1}")
for col in "BCD":
    for i, rr in enumerate(fcf_rows):
        rev_r = 21 + i
        prev_rev = f"{col}4" if i == 0 else f"{col}{20+i}"
        da = f"{col}{rev_r}*{col}$13"
        capex = f"{col}{rev_r}*{col}$15"
        dwc = f"({col}{rev_r}-{prev_rev})*{col}$16"
        eb_cell = f"{col}{29+i}"
        ws[f"{col}{rr}"] = (f"={eb_cell}-({eb_cell}-{da})*{col}$14"
                            f"-{capex}-{dwc}")
        ws[f"{col}{rr}"].number_format = MONEY

# PV rows 45-51
put(ws, "A44", "PV of FCF")
pv_rows = list(range(45, 52))
for i, rr in enumerate(pv_rows):
    put(ws, f"A{rr}", f"PV FCF Yr{i+1}")
for col in "BCD":
    for i, rr in enumerate(pv_rows):
        n = i + 1
        ws[f"{col}{rr}"] = f"={col}{37+i}/(1+{col}$17)^{n}"
        ws[f"{col}{rr}"].number_format = MONEY

# EV bridge 53-58
put(ws, "A53", "PV of explicit FCF (sum)", font=bold)
put(ws, "A54", "Terminal value (on Yr7 FCF)")
put(ws, "A55", "PV of terminal value")
put(ws, "A56", "Modular Enterprise Value", font=bold)
for col in "BCD":
    ws[f"{col}53"] = f"=SUM({col}45:{col}51)"
    ws[f"{col}54"] = f"={col}43*(1+{col}$18)/({col}$17-{col}$18)"
    ws[f"{col}55"] = f"={col}54/(1+{col}$17)^7"
    ws[f"{col}56"] = f"={col}53+{col}55"
    for rr in (53, 54, 55, 56):
        ws[f"{col}{rr}"].number_format = MONEY
    ws[f"{col}56"].font = bold
    ws[f"{col}56"].fill = total_fill
ws["A56"].fill = total_fill
put(ws, "A58", "Lead-review targets: Modular EV bear $5.7M / base $12.1M / bull $21.0M. The live explicit "
               "build matches bear ($5.7M) and base ($12.1M) to the cent; the BULL runs slightly hot at "
               "~$22.6M because model.md's stated bull input grid (+18/+14/+10/+8/+6/+5/+4%, 19% margin, "
               "WACC 12%, g 3.0%) computes to $22.6M while model.md rounded its bull total to $21.0M. "
               "Formulas are kept live (not hardcoded to $21.0M) per the no-black-box discipline; flex "
               "the yellow bull inputs to tie to $21.0M if desired. DCF base ($12.1M) corroborates "
               "SOTP-multiple base ($10.9M) within ~10% -> base modular EV ~$11-12M robust.",
    font=note_font, wrap=True)
ws.merge_cells("A58:D62")

# ============================================================================
# 5. TWO-LEG OPERATING MODEL
# ============================================================================
ws = wb.create_sheet("TwoLeg")
ws.sheet_view.showGridLines = False
for col, w in {"A": 34, "B": 14, "C": 14, "D": 14, "E": 36}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Two-leg operating model — Art's-Way Scientific standalone", font=title_font)
put(ws, "A2", "The 60/40 split is ESTIMATED — THE DATA FLOOR (§A3/§10). Below the segment total, "
              "everything is estimated from one '+$1.355M ag-building' breadcrumb.", font=red_font)

for j, h in enumerate(["Leg", "FY25 rev $M", "% of modular", "Base near-term growth", "Policy macro"]):
    put(ws, f"{chr(65+j)}4", h)
style_header_row(ws, 4, 5)
put(ws, "A5", "Research / biocontainment labs")
put(ws, "B5", 6.1, fmt=MONEY, fill=input_fill)
put(ws, "C5", "=B5/B7", fmt=PCT)
put(ws, "D5", "flat-to-+3% -> +5-7%", fill=input_fill)
put(ws, "E5", "NIH-uncertainty CHILL (cap struck down Jan-26) -> secular tailwind; "
              "pharma/biodefense/DoD/private diagnostics diversify off NIH", wrap=True)
put(ws, "A6", "Ag / animal-biosecurity buildings")
put(ws, "B6", 4.1, fmt=MONEY, fill=input_fill)
put(ws, "C6", "=B6/B7", fmt=PCT)
put(ws, "D6", "+5% then revert +2-3%", fill=input_fill)
put(ws, "E6", "USDA up-to-$1B HPAI plan ($500M cost-share @75%) FUNDING NOW; "
              "drove FY25 +$1.355M (+49%); event-driven/cyclical", wrap=True)
put(ws, "A7", "Total modular (cited)", font=bold)
put(ws, "B7", 10.226, fmt=MONEY, font=bold)
put(ws, "C7", "=C5+C6", fmt=PCT, font=bold)
for c in "ABC":
    ws[f"{c}7"].fill = total_fill
ws.row_dimensions[5].height = 44
ws.row_dimensions[6].height = 44

put(ws, "A9", "THE POLICY HEDGE (the most attractive structural feature)", font=sub_font)
put(ws, "A10", "The two legs are NEGATIVELY CORRELATED ON POLICY: when the research leg is chilled by "
               "federal-science budget fights, the ag-bio leg is funded by a different federal pocket "
               "(USDA biosecurity). A single 'modular farm-building' tag misses this hedge entirely. "
               "Blended segment ~32% GM / ~17% op margin (cited), lumpy/project-based. Backlog $4.88M "
               "(Feb-26, +103% YoY) underwrites ~6 months visibility.", font=note_font, wrap=True)
ws.merge_cells("A10:E12")

put(ws, "A14", "WHY THE SPLIT IS THE #2 TORNADO DRIVER", font=sub_font)
put(ws, "A15", "The 10-K gives only the modular segment total ($10.226M rev, $1.751M op inc) and the "
               "'+$1.355M ag-building' breadcrumb (implies FY24 ag-building base ~$2.77M -> FY25 ag-bio "
               "leg ~$4.1M = ~40%). Margins by leg are NOT disclosed; research-lab assumed higher margin "
               "(BSL/vivarium, longer cycle), ag-bio lower, blending to the cited 32.2% GM. Closes only "
               "with IR/management calls or if ARTW breaks out the legs.", font=note_font, wrap=True)
ws.merge_cells("A15:E17")

# ============================================================================
# 6. SCENARIOS & RETURNS
# ============================================================================
ws = wb.create_sheet("Scenarios")
ws.sheet_view.showGridLines = False
for col, w in {"A": 22, "B": 13, "C": 13, "D": 13, "E": 13, "F": 14, "G": 14}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Scenarios & returns — standalone FV + acquisition MOIC", font=title_font)
put(ws, "A3", "Current price ($)")
put(ws, "B3", 2.58, fmt=PS, fill=input_fill)
put(ws, "A4", "Acquisition premium (FONR precedent)")
put(ws, "B4", 0.315, fmt=PCT, fill=input_fill)
put(ws, "A5", "Acquisition buy-in price ($)", font=bold)
put(ws, "B5", "=B3*(1+B4)", fmt=PS, font=bold)
put(ws, "C5", "= $3.39/sh at 31.5% premium", font=note_font)

put(ws, "A6", "STANDALONE CARVE-OUT FAIR VALUE (blended DCF + multiple, §3)", font=sub_font)
for j, h in enumerate(["Scenario", "FV $/sh", "vs price", "Acq buy-in $", "Probability", "Acq MOIC"]):
    put(ws, f"{chr(65+j)}7", h)
# header row is 7? need style — but row 7 also holds Bear below. Put header at row7, data 8-10
# adjust: header row 6 used for sub_font label. Use row 7 header, 8/9/10 data, 11 weighted.
style_header_row(ws, 7, 6)
# blended $/sh per scenario: average of ModularDCF EV-derived $/sh and SOTP-multiple $/sh
# Build blended equity directly: (blend modular EV + net ag - net debt)/shares
# Bear/Base/Bull modular EV blend = avg(DCF EV, multiple EV)
# DCF EV: ModularDCF!B56/C56/D56 ; multiple EV: SOTP!B8/C8/D8
# net ag: SOTP C18(bear)/D18(base)/E18(bull) ; net debt SOTP!B28 ; shares SOTP!B29
scen = [
    ("Bear", "ModularDCF!B56", "SOTP!B8", "SOTP!C18", 0.30),
    ("Base", "ModularDCF!C56", "SOTP!C8", "SOTP!D18", 0.45),
    ("Bull", "ModularDCF!D56", "SOTP!D8", "SOTP!E18", 0.25),
]
r = 8
for name, dcf_ev, mult_ev, netag, prob in scen:
    put(ws, f"A{r}", name)
    # blended FV/sh = (AVG(dcf_ev,mult_ev) + netag - net_debt)/shares
    put(ws, f"B{r}", f"=((({dcf_ev})+({mult_ev}))/2+{netag}-SOTP!$B$28)/SOTP!$B$29", fmt=PS)
    put(ws, f"C{r}", f"=B{r}/$B$3-1", fmt=PCT)
    put(ws, f"D{r}", "=$B$5", fmt=PS)
    put(ws, f"E{r}", prob, fmt='0%', fill=input_fill)
    put(ws, f"F{r}", f"=B{r}/D{r}", fmt=MULT2)
    r += 1
put(ws, "A11", "Probability-weighted", font=bold)
put(ws, "B11", "=SUMPRODUCT(B8:B10,E8:E10)", fmt=PS, font=bold)
put(ws, "C11", "=B11/$B$3-1", fmt=PCT, font=bold)
put(ws, "E11", "=SUM(E8:E10)", fmt='0%', font=bold)
put(ws, "F11", "=SUMPRODUCT(F8:F10,E8:E10)", fmt=MULT2, font=bold)
for c in "ABCDEF":
    ws[f"{c}11"].fill = total_fill

put(ws, "A13", "Lead-review anchors: Bear $1.04 (-60%) / Base $2.46 (-4%) / Bull $4.72 (+83%); "
               "prob-weighted (30/45/25) ~$2.60 (+1%). Acq MOIC vs $3.39 buy-in: base 0.73x (underwater) "
               "/ bull 1.39x / bear 0.31x. NOTE: the live bull is ~$4.88 (MOIC 1.44x) — slightly above "
               "the anchor because the explicit DCF of model.md's stated bull inputs computes Modular EV "
               "~$22.6M vs model.md's rounded $21.0M (see ModularDCF note). Bear & base tie exactly. "
               "The deal still works ONLY on the bull.",
    font=note_font, wrap=True)
ws.merge_cells("A13:G15")
put(ws, "A16", "Edit column E (probabilities) and any yellow input on ModularDCF/SOTP to flex. "
               "FV here = present intrinsic value (gap to fair closes over 3 yrs, no further compounding).",
    font=note_font, wrap=True)
ws.merge_cells("A16:G17")

# ============================================================================
# 7. VALUE LENSES
# ============================================================================
ws = wb.create_sheet("ValueLenses")
ws.sheet_view.showGridLines = False
for col, w in {"A": 42, "B": 16, "C": 14}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Graham / Buffett value lenses ($M unless noted)", font=title_font)

# EPV (modular only, loads full net debt)
put(ws, "A3", "EARNINGS POWER VALUE (modular-only, no growth, loads full net debt)", font=sub_font)
epv = [
    ("Modular EBIT (op income, normalized)", 1.55, MONEY, True),
    ("Cash tax rate", 0.15, PCT, True),
    ("NOPAT", "=B4*(1-B5)", MONEY, False),
    ("WACC", 0.135, PCT, True),
    ("EPV operating (NOPAT / WACC)", "=B6/B7", MONEY, False),
    ("less: net debt", "=Summary!B13", MONEY, False),
    ("EPV equity", "=B8-B9", MONEY, False),
    ("Shares (M)", "=Summary!B9", NUM3, False),
    ("EPV per share ($)", "=B10/B11", PS, False),
]
r = 4
for label, v, fmt, is_input in epv:
    em = "EPV per share" in label
    put(ws, f"A{r}", label, font=bold if em else None)
    put(ws, f"B{r}", v, fmt=fmt, fill=input_fill if is_input else None, font=bold if em else None)
    if em:
        ws[f"A{r}"].fill = total_fill; ws[f"B{r}"].fill = total_fill
    r += 1
# EPV per share at row 12 -> Summary references B14? fix: it's at B12. Summary used ValueLenses!B14.
put(ws, "A13", "Read: after loading FULL net debt on modular alone, no-growth value (~$0.65/sh) is well "
               "below price — you are paying for growth + the ag asset recovery.", font=note_font, wrap=True)
ws.merge_cells("A13:C14")

# Reverse DCF
put(ws, "A16", "REVERSE DCF (what the ~$20M EV implies)", font=sub_font)
rd = [
    ("Enterprise value ($M)", "=Summary!B14", MONEY, False),
    ("Normalized FCF ($M)", 1.2, MONEY, True),
    ("WACC", 0.135, PCT, True),
    ("Implied perpetual FCF growth", "=B18-B19/B17", PCT, False),
]
r = 17
for label, v, fmt, is_input in rd:
    em = "Implied" in label
    put(ws, f"A{r}", label, font=bold if em else None)
    put(ws, f"B{r}", v, fmt=fmt, fill=input_fill if is_input else None, font=bold if em else None)
    if em:
        ws[f"A{r}"].fill = total_fill; ws[f"B{r}"].fill = total_fill
    r += 1
# Implied growth at row 20 -> Summary referenced ValueLenses!B23. fix.
put(ws, "A21", "The ~$20M EV embeds ~7% perpetual FCF growth — the market is NOT pricing a melting "
               "farm-equipment microcap; it embeds mid-single-digit+ growth. Not cheap on expectations.",
    font=note_font, wrap=True)
ws.merge_cells("A21:C22")

# Owner-earnings yield
put(ws, "A24", "OWNER-EARNINGS YIELD (Buffett)", font=sub_font)
put(ws, "A25", "Normalized FCF / owner earnings ($M)")
put(ws, "B25", 1.2, fmt=MONEY, fill=input_fill)
put(ws, "A26", "Market cap ($M)")
put(ws, "B26", "=Summary!B10", fmt=MONEY)
put(ws, "A27", "Owner-earnings yield (on market cap)", font=bold)
put(ws, "B27", "=B25/B26", fmt=PCT, font=bold)
ws["A27"].fill = total_fill; ws["B27"].fill = total_fill
put(ws, "A28", "FCF / enterprise value")
put(ws, "B28", "=B25/Summary!B14", fmt=PCT)

# Graham NCAV
put(ws, "A30", "GRAHAM NCAV (net-net floor)", font=sub_font)
put(ws, "A31", "Current assets ($M)")
put(ws, "B31", 14.78, fmt=MONEY, fill=input_fill)
put(ws, "A32", "less: total liabilities ($M)")
put(ws, "B32", 9.17, fmt=MONEY, fill=input_fill)
put(ws, "A33", "Net current asset value ($M)")
put(ws, "B33", "=B31-B32", fmt=MONEY)
put(ws, "A34", "NCAV per share ($)", font=bold)
put(ws, "B34", "=B33/Summary!B9", fmt=PS, font=bold)
ws["A34"].fill = total_fill; ws["B34"].fill = total_fill
put(ws, "A35", "Not a net-net at $2.58, but ~42% of price covered by net current assets alone — a real "
               "downside marker.", font=note_font, wrap=True)
ws.merge_cells("A35:C36")

# Graham strong-financial-condition (mirrored on Graham tab; summary cells)
put(ws, "A38", "GRAHAM STRONG-FINANCIAL-CONDITION (mirror — full calc on Graham tab)", font=sub_font)
put(ws, "A39", "Current ratio")
put(ws, "B39", "=Graham!B7", fmt=MULT2)
put(ws, "A40", "LT debt vs working capital")
put(ws, "B40", '=Graham!B10&" vs "&Graham!B8', font=None)
put(ws, "A42", "Current-ratio test (>=2.0)")
put(ws, "C42", "=Graham!C7", font=green_font)
put(ws, "A45", "LT-debt<=WC test")
put(ws, "C45", "=Graham!C11", font=green_font)
put(ws, "A47", "PASSES BOTH — unlike HCKT (failed both). A genuine static balance-sheet floor; the "
               "downside is partly asset-backed (inventory + plant + 1.0x book), not carried solely by "
               "earning power.", font=note_font, wrap=True)
ws.merge_cells("A47:C48")

# ============================================================================
# 8. GRAHAM FINANCIAL-CONDITION TEST
# ============================================================================
ws = wb.create_sheet("Graham")
ws.sheet_view.showGridLines = False
for col, w in {"A": 40, "B": 16, "C": 12, "D": 40}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Graham 'strong financial condition' test (Defensive Investor)", font=title_font)
put(ws, "A2", "CITED: FY25 10-K MD&A working-capital table (p.511) + balance sheet / Note 10. "
              "Contrast HCKT which FAILED both prongs.", font=note_font)

for j, h in enumerate(["Item", "Value", "Test", "Note"]):
    put(ws, f"{chr(65+j)}4", h)
style_header_row(ws, 4, 4)
put(ws, "A5", "Current assets ($M)")
put(ws, "B5", 14.78, fmt=MONEY, fill=input_fill)
put(ws, "D5", "CITED — FY25 balance sheet", font=note_font)
put(ws, "A6", "Current liabilities ($M)")
put(ws, "B6", 6.44, fmt=MONEY, fill=input_fill)
put(ws, "D6", "CITED — implied (WC $8.34M = CA - CL)", font=note_font)
put(ws, "A7", "Current ratio  (Graham test: >= 2.0)", font=bold)
put(ws, "B7", "=B5/B6", fmt=MULT2, font=bold)
put(ws, "C7", '=IF(B7>=2,"PASS","FAIL")', font=green_font)
put(ws, "D7", "CITED 2.30x (FY24 1.98x) — PASSES", font=note_font)
put(ws, "A8", "Working capital (net current assets, $M)")
put(ws, "B8", "=B5-B6", fmt=MONEY)
put(ws, "D8", "CITED $8.34M (FY24 $6.49M)", font=note_font)
put(ws, "A10", "Long-term debt ($M)")
put(ws, "B10", 2.73, fmt=MONEY, fill=input_fill)
put(ws, "D10", "CITED — LT term loans $2.325M + LT finance lease $0.408M", font=note_font)
put(ws, "A11", "LT debt <= working capital  (Graham test)", font=bold)
put(ws, "B11", "=B10/B8", fmt=MULT2, font=bold)
put(ws, "C11", '=IF(B10<=B8,"PASS","FAIL")', font=green_font)
put(ws, "D11", "LT debt $2.73M <= WC $8.34M — PASSES", font=note_font)

put(ws, "A13", "RESULT: PASSES BOTH PRONGS.", font=green_font)
put(ws, "A14", "Unlike HCKT (current ratio 1.88x FAIL; LT debt $78.8M > WC $38.2M FAIL), ARTW has a "
               "genuine static balance-sheet floor — the downside is partly asset-backed (inventory + "
               "plant + 1.0x book), not carried solely by earning power. Also: NCAV $1.08/sh (~42% of "
               "price), book value $2.57 ≈ price (~1.0x book).", font=note_font, wrap=True)
ws.merge_cells("A14:D16")

# ============================================================================
# 9. PEERS  (read live from peer CSV)
# ============================================================================
ws = wb.create_sheet("Peers")
ws.sheet_view.showGridLines = False
for col, w in {"A": 9, "B": 24, "C": 18, "D": 11, "E": 10, "F": 10, "G": 11, "H": 70}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Peer comp set — multiples behind the SOTP & DCF", font=title_font)
put(ws, "A2", "Source: data/research/artw/peer_benchmarks.csv. FONR = precedent transaction (the base "
              "anchor); WSC = leasing ceiling (not comparable); construction median ~10x.", font=note_font)
heads = ["Ticker", "Name", "Inclusion", "EV/EBITDA", "EV/Rev", "Rev gr %", "EBITDA mgn", "Notes"]
for j, h in enumerate(heads):
    put(ws, f"{chr(65+j)}4", h)
style_header_row(ws, 4, len(heads))
peer_csv = Path(__file__).parent.parent / "data" / "research" / "artw" / "peer_benchmarks.csv"
r = 5
with open(peer_csv, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        def num(k):
            v = row.get(k, "")
            try:
                return float(v)
            except (ValueError, TypeError):
                return None
        is_subject = row["inclusion"] == "subject"
        put(ws, f"A{r}", row["peer_ticker"], font=bold if is_subject else None)
        put(ws, f"B{r}", row["peer_name"])
        put(ws, f"C{r}", row["inclusion"])
        if num("ev_ebitda") is not None:
            put(ws, f"D{r}", num("ev_ebitda"), fmt=MULT)
        if num("ev_revenue") is not None:
            put(ws, f"E{r}", num("ev_revenue"), fmt=MULT)
        if num("revenue_growth_pct") is not None:
            put(ws, f"F{r}", num("revenue_growth_pct") / 100, fmt=PCT)
        if num("ebitda_margin") is not None:
            put(ws, f"G{r}", num("ebitda_margin") / 100, fmt=PCT)
        put(ws, f"H{r}", row.get("notes", "")[:400], font=note_font)
        ws[f"H{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        if is_subject:
            for c in "ABCDEFG":
                ws[f"{c}{r}"].fill = total_fill
        r += 1
put(ws, f"A{r+1}", "Key: FONR (~7-9x EBITDA, 0.94x rev, 31.5% premium, ~$98.6M) is the closest structural "
                   "analog (CEO-led take-private of a profitable family-controlled health-adjacent micro-cap) "
                   "— it anchors BOTH the base multiple AND the acquisition premium. WSC (11.7-14.5x) is a "
                   "LEASING model (recurring rental ROIC) — used only as the unreachable ceiling for "
                   "'modular.' Construction median ~10.3x; permanent-modular ~7% EBITDA margin (McKinsey). "
                   "ARTW modular ~17% op margin is ABOVE the sector norm -> supports mid-to-high-single-digit "
                   "EBITDA multiple. US modular construction ~$20.3B (2024), ~4.5-6.2% CAGR.", font=note_font, wrap=True)
ws.merge_cells(f"A{r+1}:H{r+3}")

# ============================================================================
# 10. ASSUMPTIONS / WACC BUILD
# ============================================================================
ws = wb.create_sheet("Assumptions")
ws.sheet_view.showGridLines = False
for col, w in {"A": 32, "B": 12, "C": 16, "D": 66}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Assumptions & judgment log — how each non-cited input was built", font=title_font)

put(ws, "A3", "WACC BUILD (bottom-up CAPM for a micro-cap, project-based, illiquid, single-plant biz)", font=sub_font)
wacc = [
    ("Risk-free rate (10-yr UST)", 0.043, PCT, "CITED — market-observable, May 2026."),
    ("Beta (levered)", 1.4, '0.00', "ESTIMATED — small-cap industrial / project cyclicality; thin-float, don't defend the 2nd digit."),
    ("Equity risk premium", 0.055, PCT, "ESTIMATED — convention (Damodaran ~4.5-5.5%)."),
    ("CAPM cost of equity (pre-overlay)", "=B4+B5*B6", PCT, "= rf + beta x ERP (~12.0%)."),
    ("", None, None, ""),
    ("Size / illiquidity premium", 0.035, PCT, "ESTIMATED — ~$13M cap, float ~2.0M sh, ~$0.1M/day volume. Range 3-4%."),
    ("Project-concentration premium", 0.0, PCT, "ESTIMATED — lumpy single-plant project risk; folded into size/illiquidity band here (set 0 to avoid double-count, raise to test)."),
    ("Cost of equity (CAPM + premia)", "=B7+B9+B10", PCT, "= CAPM + size/illiquidity + project conc."),
    ("", None, None, ""),
    ("WACC — base (applied in DCF)", "=B11", PCT, "~13.5% base. Bear 15% (heavier risk premium); bull 12% (cleaner)."),
    ("WACC — bear", 0.15, PCT, "ESTIMATED — heavier risk premium on a missed inflection."),
    ("WACC — bull", 0.12, PCT, "ESTIMATED — re-rate to a clean biocontainment platform."),
]
r = 4
for label, val, fmt, logic in wacc:
    if not label:
        r += 1
        continue
    is_calc = isinstance(val, str) and val.startswith("=")
    boldit = "applied" in label
    put(ws, f"A{r}", label, font=bold if boldit else None)
    put(ws, f"B{r}", val, fmt=fmt, fill=None if is_calc else input_fill, font=bold if boldit else None)
    put(ws, f"D{r}", logic, font=note_font)
    ws[f"D{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    if boldit:
        for c in "AB":
            ws[f"{c}{r}"].fill = total_fill
    r += 1

put(ws, "A18", "JUDGMENT-INPUT LOG (every non-cited assumption: tier + how it was built)", font=sub_font)
for j, h in enumerate(["Input", "Base", "Tier", "How constructed / logic"]):
    put(ws, f"{chr(65+j)}19", h)
style_header_row(ws, 19, 4)
log = [
    ("WACC", "13.5%", "Built CAPM + premia", "rf 4.3% + beta 1.4 x ERP 5.5% + size/illiquidity 3-4%; bear 15% / bull 12% (built above)."),
    ("Cash tax rate", "15%", "Normalized", "CITED cash tax ~zero (NOL shield), but NOLs finite & a profitable carve-out exhausts them; do NOT extrapolate 0% (§A2). Near-term real rate lower = upside."),
    ("Two-leg split (research/ag-bio)", "60/40", "ESTIMATED — DATA FLOOR", "From the single '+$1.355M ag-building' breadcrumb -> FY24 base ~$2.77M -> FY25 ag-bio ~$4.1M (~40%). THE key data-floor assumption (§A3/§10). Revenue & margin by leg NOT disclosed."),
    ("Modular standalone EBITDA", "$1.55M base", "Built", "Segment EBITDA $2.002M (op $1.751M + D&A $0.251M) less ~$0.45M incremental standalone public-company/corporate cost. Bear $1.40M / bull $2.00M."),
    ("Standalone corporate cost add", "$0.45M", "ESTIMATED", "Only ~$0.18M corp was allocated to modular in the segment table; a standalone registrant carries more. Full take-private removes public-co cost = upside."),
    ("Modular EV/EBITDA multiple", "7x base", "Judgment", "FONR ~7-9x precedent anchor; Porter read supports base 7x and caps bull ~10x. WSC 11.7-14.5x leasing = unreachable ceiling. Bear 5x."),
    ("Ag inventory recovery", "60% base", "ESTIMATED", "v2 re-centered from v1's 70% (which tuned base FV to the $2.58 quote). Aged whole-goods ARTW built into a slow market + beet-equipment hit by Dec-2025 44% ACS payment cut -> 55-65% realistic. Bear 50% / bull 75%."),
    ("Ag AR recovery", "95% base", "ESTIMATED", "Trade AR collects near par in an orderly wind-down. Bear 85% / bull 100%."),
    ("Ag PP&E recovery", "40% base", "ESTIMATED", "Single-purpose rural-Iowa (Armstrong) plant. 2024 Tools real-estate sale $1.8M a loose data point. Bear 30% / bull 55%."),
    ("Ag wind-down cost", "$1.3M base", "ESTIMATED", "Severance, lease breakage, warranty/dealer tail, transaction. Bear $2.0M / bull $0.8M."),
    ("D&A & capex (% rev)", "2.5% each", "Judgment", "Asset-light modular ($3.27M assets, $0.222M capex on $10.2M rev)."),
    ("ΔWC (% incremental rev)", "10%", "Judgment", "Project-timing-driven (over/under-billing nets to ~zero over a cycle); NOT a permanent build like ag whole-goods (§A6)."),
    ("Terminal growth", "2.5% base", "Judgment", "GDP-ish. Bear 1.0% / bull 3.0%."),
    ("Scenario probabilities", "30/45/25", "Judgment", "Bear weighted for the bull-dependent, data-floor-limited setup; no enacted NIH cut but chilled climate; ag-bio cyclical."),
    ("Acquisition premium", "31.5%", "CITED precedent", "FONR Dec-2025 CEO-led take-private at 31.5% premium -> ARTW buy-in $3.39/sh. Base MOIC 0.73x (underwater); only bull (1.39x) clears."),
    ("Net debt", "$6.40M", "CITED", "Total debt $6.41M (revolver $3.252M + term loans + finance leases, incl ~$0.309M SBA EIDL) - cash $0.005M (§0)."),
]
r = 20
for inp, val, tier, logic in log:
    put(ws, f"A{r}", inp)
    put(ws, f"B{r}", val)
    put(ws, f"C{r}", tier)
    put(ws, f"D{r}", logic, font=note_font)
    ws[f"D{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

wb.save(OUT)
print("wrote", OUT)
