"""
Build a live, formula-driven Excel model for HCKT (The Hackett Group) from the
v2 deep-flow build in data/research/hckt/model.md.

Every output cell is a formula referencing editable input cells, so the reviewer
can flex WACC / growth / margins / segment multiples / scenario probabilities and
watch the DCF, SOTP, scenario IRRs, and the Graham/Buffett value lenses recompute.

Output: data/research/hckt/HCKT_model.xlsx
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule

OUT = Path(__file__).parent.parent / "data" / "research" / "hckt" / "HCKT_model.xlsx"

NAVY = "0A2540"
RED = "C8102E"
TINT = "FEF7F7"
GREY = "6B7280"
LIGHT = "F1F3F5"

hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=NAVY)
sub_font = Font(name="Calibri", bold=True, color=NAVY, size=12)
title_font = Font(name="Calibri", bold=True, color=NAVY, size=15)
red_font = Font(name="Calibri", bold=True, color=RED, size=11)
bold = Font(name="Calibri", bold=True, size=11)
note_font = Font(name="Calibri", italic=True, color=GREY, size=9)
total_fill = PatternFill("solid", fgColor=TINT)
input_fill = PatternFill("solid", fgColor="FFFDE7")  # pale yellow = editable input
thin = Side(style="thin", color="D9DCE1")
border = Border(bottom=thin)

MONEY = '"$"#,##0.0'
PS = '"$"#,##0.00'
PCT = '0.0%'
MULT = '0.0"x"'

wb = Workbook()


def style_header_row(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")


def put(ws, addr, value, *, fmt=None, font=None, fill=None, align=None):
    cell = ws[addr]
    cell.value = value
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = Alignment(horizontal=align)
    return cell


# ----------------------------------------------------------------------------
# DASHBOARD
# ----------------------------------------------------------------------------
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False
for col, w in {"A": 34, "B": 16, "C": 16, "D": 16, "E": 22}.items():
    ws.column_dimensions[col].width = w

put(ws, "A1", "The Hackett Group (HCKT) — Deep-Flow Valuation Model", font=title_font)
put(ws, "A2", "v2 (lead-reviewed) · 2026-05-29 · $ in millions unless noted · yellow = editable input",
    font=note_font)

put(ws, "A4", "VERDICT", font=red_font)
put(ws, "A5", "Mispriced-but-show-me: the AI-disintermediation bear FAILS on the filings "
              "(AI is raising delivery margins); the bear survives only on near-term revenue. "
              "Catalyst = the Q3-FY26 inflection. No asset floor — size for a real bear.")
ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A5:E5")
ws.row_dimensions[5].height = 46

put(ws, "A7", "KEY FACTS", font=sub_font)
facts = [
    ("Current price ($)", "=Scenarios!B3", PS),
    ("Shares outstanding (M)", 25.18, '#,##0.00'),
    ("Market cap ($M)", "=B8*B9", MONEY),
    ("Net financial debt ($M)", 72.7, MONEY),
    ("Enterprise value ($M)", "=B10+B11", MONEY),
    ("EV / adj. EBITDA (~$60M)", "=B12/60", MULT),
    ("Free-cash-flow yield (on mkt cap)", "=32.4/B10", PCT),
    ("Dividend yield", 0.042, PCT),
]
r = 8
for label, val, fmt in facts:
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", val, fmt=fmt)
    r += 1

put(ws, "A17", "HEADLINE OUTPUTS  (live from other tabs)", font=sub_font)
outs = [
    ("Base fair value / share — DCF", "=DCF!C50", PS),
    ("Base fair value / share — SOTP", "=SOTP!G17", PS),
    ("Base fair value / share — blended", "=Scenarios!D8", PS),
    ("Probability-weighted 3-yr FV / share", "=Scenarios!D10", PS),
    ("Probability-weighted price IRR (/yr)", "=Scenarios!F10", PCT),
    ("Probability-weighted total IRR (+div)", "=Scenarios!G10", PCT),
]
r = 18
for label, val, fmt in outs:
    put(ws, f"A{r}", label, font=bold if "blended" in label or "total" in label else None)
    put(ws, f"B{r}", val, fmt=fmt, font=bold if "blended" in label or "total" in label else None)
    if "total" in label or "blended" in label:
        ws[f"A{r}"].fill = total_fill
        ws[f"B{r}"].fill = total_fill
    r += 1

put(ws, "A25", "GRAHAM / BUFFETT VALUE LENSES  (live from ValueLenses tab)", font=sub_font)
lenses = [
    ("Earnings Power Value (no growth), charging SBC", "=ValueLenses!B16", PS),
    ("Earnings Power Value (no growth), not charging SBC", "=ValueLenses!C16", PS),
    ("Reverse DCF — perpetual FCF growth priced in", "=ValueLenses!B27", PCT),
    ("Owner-earnings yield (on market cap)", "=ValueLenses!B33", PCT),
    ("Graham asset floor", "NONE — net debt + goodwill", None),
]
r = 26
for label, val, fmt in lenses:
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", val, fmt=fmt)
    if fmt is None:
        ws[f"B{r}"].font = red_font
    r += 1
put(ws, "A31", "Graham strong-financial-condition (curr-ratio / LT-debt≤WC)")
put(ws, "B31", '=ValueLenses!C46&" / "&ValueLenses!C49', font=red_font)

put(ws, "A33", "Read: at the current price you are paying ~the no-growth Earnings Power Value — the "
               "recovery/re-rate is a near-free option, but there is no Graham margin of safety below it, "
               "and no asset floor under the bear.", font=note_font)
ws["A33"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A33:E34")

put(ws, "A36", "TABS: Operating · DCF · SOTP · Scenarios · ValueLenses · Sensitivity", font=note_font)

# ----------------------------------------------------------------------------
# OPERATING MODEL
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Operating")
ws.sheet_view.showGridLines = False
for col, w in {"A": 30, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Operating model — actuals & segment economics ($M)", font=title_font)

put(ws, "A3", "Historicals")
for j, h in enumerate(["", "FY23", "FY24", "FY25", "Q1 FY26"]):
    put(ws, f"{chr(65+j)}3", h)
style_header_row(ws, 3, 5)
rows = [
    ("Total revenue", [296.6, 313.9, 305.6, 68.8], MONEY),
    ("  growth YoY", [None, "=C4/B4-1", "=D4/C4-1", -0.116], PCT),
    ("GAAP operating income", [49.3, 44.6, 23.5, 8.9], MONEY),
    ("  GAAP operating margin", ["=B6/B4", "=C6/C4", "=D6/D4", "=E6/E4"], PCT),
    ("Adjusted EBITDA (mgmt)", [None, None, 60.0, 13.8], MONEY),
    ("Net income", [34.2, 29.6, 12.9, 4.3], MONEY),
    ("Operating cash flow", [37.4, 47.7, 40.3, -5.1], MONEY),
    ("Free cash flow (OCF − capex)", [33.3, 43.6, 32.4, None], MONEY),
    ("Total stock-based comp (non-cash)", [10.7, 19.5, 30.7, 2.4], MONEY),
    ("  of which one-time award", [0, 5.7, 16.8, None], MONEY),
    ("Cash income taxes paid", [13.3, 11.6, 13.1, 0.6], MONEY),
    ("Consultant headcount", [None, 1332, 1301, 1247], '#,##0'),
]
r = 4
for label, vals, fmt in rows:
    put(ws, f"A{r}", label)
    for j, v in enumerate(vals):
        if v is not None:
            put(ws, f"{chr(66+j)}{r}", v, fmt=fmt)
    r += 1

put(ws, "A18", "FY25 segment economics", font=sub_font)
for j, h in enumerate(["Segment", "Revenue", "% of rev", "Contribution", "Margin"]):
    put(ws, f"{chr(65+j)}19", h)
style_header_row(ws, 19, 5)
seg = [
    ("Global S&BT (advisory + benchmark IP)", 169.6, 50.8),
    ("Oracle Solutions (implementation)", 72.7, 12.4),
    ("SAP Solutions (implementation + license)", 63.4, 20.4),
]
r = 20
for name, rev, contr in seg:
    put(ws, f"A{r}", name)
    put(ws, f"B{r}", rev, fmt=MONEY)
    put(ws, f"C{r}", f"=B{r}/B23", fmt=PCT)
    put(ws, f"D{r}", contr, fmt=MONEY)
    put(ws, f"E{r}", f"=D{r}/B{r}", fmt=PCT)
    r += 1
put(ws, "A23", "Total", font=bold)
put(ws, "B23", "=SUM(B20:B22)", fmt=MONEY, font=bold)
put(ws, "D23", "=SUM(D20:D22)", fmt=MONEY, font=bold)
ws["E23"] = "=D23/B23"; ws["E23"].number_format = PCT; ws["E23"].font = bold
for c in "ABCDE":
    ws[f"{c}23"].fill = total_fill

put(ws, "A25", "Key operating insight: earnings inflect BEFORE revenue. Q1 FY26 revenue −11.6% YoY "
               "but GAAP operating income +103% (margin 5.7%→13.0%), as the one-time stock award rolls "
               "off and AI lifts delivery margins (+500bps). ~24% of revenue is recurring.", font=note_font)
ws["A25"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A25:F26")

# ----------------------------------------------------------------------------
# DCF
# ----------------------------------------------------------------------------
ws = wb.create_sheet("DCF")
ws.sheet_view.showGridLines = False
for col, w in {"A": 30, "B": 13, "C": 13, "D": 13}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Discounted cash flow — 3 scenarios ($M)", font=title_font)
put(ws, "A2", "Cash taxes from the cash-flow statement (~26% normalized), not the 40.7% book rate.",
    font=note_font)

for j, h in enumerate(["Driver / Year", "Bear", "Base", "Bull"]):
    put(ws, f"{chr(65+j)}3", h)
style_header_row(ws, 3, 4)

inputs = [
    ("Base-year revenue (FY25)", [305.6, 305.6, 305.6], MONEY),
    ("Yr1 revenue growth", [-0.05, -0.04, -0.02], PCT),
    ("Yr2 revenue growth", [-0.02, 0.02, 0.06], PCT),
    ("Yr3 revenue growth", [0.0, 0.04, 0.08], PCT),
    ("Yr4 revenue growth", [0.01, 0.04, 0.06], PCT),
    ("Yr5 revenue growth", [0.01, 0.04, 0.05], PCT),
    ("Adj. EBITDA margin", [0.18, 0.20, 0.22], PCT),
    ("D&A", [5.2, 5.2, 5.2], MONEY),
    ("Cash tax rate", [0.28, 0.26, 0.24], PCT),
    ("Capex", [6.5, 6.0, 6.0], MONEY),
    ("Working-capital use", [1.5, 1.5, 1.0], MONEY),
    ("WACC", [0.115, 0.105, 0.095], PCT),
    ("Terminal growth", [0.0, 0.01, 0.03], PCT),
    ("Net debt", [72.7, 72.7, 72.7], MONEY),
    ("Shares (M)", [25.18, 25.18, 25.18], '#,##0.00'),
]
r = 4
for label, vals, fmt in inputs:
    put(ws, f"A{r}", label)
    for j, v in enumerate(vals):
        put(ws, f"{chr(66+j)}{r}", v, fmt=fmt, fill=input_fill)
    r += 1
# rows: rev base=4, g1=5,g2=6,g3=7,g4=8,g5=9, margin=10, D&A=11, tax=12, capex=13, wc=14, wacc=15, termg=16, netdebt=17, shares=18

put(ws, "A20", "PROJECTION", font=sub_font)
put(ws, "A21", "Revenue Yr1")
put(ws, "A22", "Revenue Yr2"); put(ws, "A23", "Revenue Yr3")
put(ws, "A24", "Revenue Yr4"); put(ws, "A25", "Revenue Yr5")
for col in "BCD":
    ws[f"{col}21"] = f"={col}4*(1+{col}5)"
    ws[f"{col}22"] = f"={col}21*(1+{col}6)"
    ws[f"{col}23"] = f"={col}22*(1+{col}7)"
    ws[f"{col}24"] = f"={col}23*(1+{col}8)"
    ws[f"{col}25"] = f"={col}24*(1+{col}9)"
    for rr in range(21, 26):
        ws[f"{col}{rr}"].number_format = MONEY

put(ws, "A27", "EBITDA Yr1"); put(ws, "A28", "EBITDA Yr2"); put(ws, "A29", "EBITDA Yr3")
put(ws, "A30", "EBITDA Yr4"); put(ws, "A31", "EBITDA Yr5")
for col in "BCD":
    for k, rr in enumerate(range(27, 32)):
        ws[f"{col}{rr}"] = f"={col}{21+k}*{col}$10"
        ws[f"{col}{rr}"].number_format = MONEY

put(ws, "A33", "FCF Yr1"); put(ws, "A34", "FCF Yr2"); put(ws, "A35", "FCF Yr3")
put(ws, "A36", "FCF Yr4"); put(ws, "A37", "FCF Yr5")
for col in "BCD":
    for k, rr in enumerate(range(33, 38)):
        eb = 27 + k
        # FCF = EBITDA - (EBITDA - D&A)*tax - capex - WC
        ws[f"{col}{rr}"] = f"={col}{eb}-({col}{eb}-{col}$11)*{col}$12-{col}$13-{col}$14"
        ws[f"{col}{rr}"].number_format = MONEY

put(ws, "A39", "PV of FCF Yr1"); put(ws, "A40", "PV Yr2"); put(ws, "A41", "PV Yr3")
put(ws, "A42", "PV Yr4"); put(ws, "A43", "PV Yr5")
for col in "BCD":
    for k, rr in enumerate(range(39, 44)):
        n = k + 1
        ws[f"{col}{rr}"] = f"={col}{33+k}/(1+{col}$15)^{n}"
        ws[f"{col}{rr}"].number_format = MONEY

put(ws, "A45", "PV of explicit FCF (sum)", font=bold)
put(ws, "A46", "Terminal value (Yr5)")
put(ws, "A47", "PV of terminal value")
put(ws, "A48", "Enterprise value", font=bold)
put(ws, "A49", "Equity value (− net debt)", font=bold)
put(ws, "A50", "Value per share ($)", font=bold)
for col in "BCD":
    ws[f"{col}45"] = f"=SUM({col}39:{col}43)"
    ws[f"{col}46"] = f"={col}37*(1+{col}$16)/({col}$15-{col}$16)"
    ws[f"{col}47"] = f"={col}46/(1+{col}$15)^5"
    ws[f"{col}48"] = f"={col}45+{col}47"
    ws[f"{col}49"] = f"={col}48-{col}$17"
    ws[f"{col}50"] = f"={col}49/{col}$18"
    for rr in (45, 46, 47, 48, 49):
        ws[f"{col}{rr}"].number_format = MONEY
    ws[f"{col}50"].number_format = PS
    ws[f"{col}50"].font = bold
    ws[f"{col}50"].fill = total_fill
ws["A50"].fill = total_fill

# ----------------------------------------------------------------------------
# SOTP
# ----------------------------------------------------------------------------
ws = wb.create_sheet("SOTP")
ws.sheet_view.showGridLines = False
for col, w in {"A": 34, "B": 14, "C": 10, "D": 10, "E": 10, "F": 12, "G": 12, "H": 12}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Sum-of-the-parts ($M) — segment contribution × peer multiple", font=title_font)

for j, h in enumerate(["Segment", "FY25 contrib.", "Bear x", "Base x", "Bull x",
                       "Bear EV", "Base EV", "Bull EV"]):
    put(ws, f"{chr(65+j)}3", h)
style_header_row(ws, 3, 8)
segs = [
    ("Global S&BT (advisory + benchmark IP)", 50.8, 6, 9, 12),
    ("Oracle Solutions (implementation)", 12.4, 5, 6.5, 8),
    ("SAP Solutions (implementation + license)", 20.4, 5, 6, 7.5),
]
r = 4
for name, contr, b, ba, bu in segs:
    put(ws, f"A{r}", name)
    put(ws, f"B{r}", contr, fmt=MONEY)
    put(ws, f"C{r}", b, fmt=MULT, fill=input_fill)
    put(ws, f"D{r}", ba, fmt=MULT, fill=input_fill)
    put(ws, f"E{r}", bu, fmt=MULT, fill=input_fill)
    put(ws, f"F{r}", f"=B{r}*C{r}", fmt=MONEY)
    put(ws, f"G{r}", f"=B{r}*D{r}", fmt=MONEY)
    put(ws, f"H{r}", f"=B{r}*E{r}", fmt=MONEY)
    r += 1
put(ws, "A7", "Total segment EV", font=bold)
for col in "FGH":
    ws[f"{col}7"] = f"=SUM({col}4:{col}6)"
    ws[f"{col}7"].number_format = MONEY
    ws[f"{col}7"].font = bold

put(ws, "A9", "Recurring cash corporate (G&A $20.5M + recurring SBC $10.9M)")
put(ws, "B9", 31.5, fmt=MONEY, fill=input_fill)
put(ws, "A10", "Corporate capitalization multiple")
put(ws, "C10", 6.5, fmt=MULT, fill=input_fill)
put(ws, "D10", 7.9, fmt=MULT, fill=input_fill)
put(ws, "E10", 9.0, fmt=MULT, fill=input_fill)
put(ws, "A11", "Capitalized corporate (−)")
put(ws, "F11", "=B9*C10", fmt=MONEY)
put(ws, "G11", "=B9*D10", fmt=MONEY)
put(ws, "H11", "=B9*E10", fmt=MONEY)
put(ws, "A12", "AI-platform optionality")
put(ws, "F12", 0, fmt=MONEY, fill=input_fill)
put(ws, "G12", 25, fmt=MONEY, fill=input_fill)
put(ws, "H12", 60, fmt=MONEY, fill=input_fill)
put(ws, "A13", "Operating EV", font=bold)
for col in "FGH":
    ws[f"{col}13"] = f"={col}7-{col}11+{col}12"
    ws[f"{col}13"].number_format = MONEY
    ws[f"{col}13"].font = bold
put(ws, "A14", "Net debt")
put(ws, "B14", 72.7, fmt=MONEY, fill=input_fill)
put(ws, "A15", "Equity value", font=bold)
for col in "FGH":
    ws[f"{col}15"] = f"={col}13-B14"
    ws[f"{col}15"].number_format = MONEY
    ws[f"{col}15"].font = bold
put(ws, "A16", "Shares (M)")
put(ws, "B16", 25.18, fmt='#,##0.00', fill=input_fill)
put(ws, "A17", "Value per share ($)", font=bold)
for col in "FGH":
    ws[f"{col}17"] = f"={col}15/B16"
    ws[f"{col}17"].number_format = PS
    ws[f"{col}17"].font = bold
    ws[f"{col}17"].fill = total_fill
ws["A17"].fill = total_fill

# ----------------------------------------------------------------------------
# SCENARIOS & IRR
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Scenarios")
ws.sheet_view.showGridLines = False
for col, w in {"A": 24, "B": 13, "C": 13, "D": 15, "E": 13, "F": 13, "G": 16}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Scenarios & probability-weighted return", font=title_font)
put(ws, "A3", "Current price ($)")
put(ws, "B3", 11.61, fmt=PS, fill=input_fill)
put(ws, "A4", "Dividend yield")
put(ws, "B4", 0.042, fmt=PCT, fill=input_fill)

for j, h in enumerate(["Scenario", "DCF $/sh", "SOTP $/sh", "Blended 3-yr FV",
                       "Probability", "Price IRR/yr", "Total IRR (+div)"]):
    put(ws, f"{chr(65+j)}6", h)
style_header_row(ws, 6, 7)
# Bear r7, Base r8, Bull r9, Weighted r10
mapping = [("Bear", "B50", "F17", 0.45), ("Base", "C50", "G17", 0.42), ("Bull", "D50", "H17", 0.13)]
r = 7
for name, dcf, sotp, prob in mapping:
    put(ws, f"A{r}", name)
    put(ws, f"B{r}", f"=DCF!{dcf}", fmt=PS)
    put(ws, f"C{r}", f"=SOTP!{sotp}", fmt=PS)
    put(ws, f"D{r}", f"=AVERAGE(B{r}:C{r})", fmt=PS)
    put(ws, f"E{r}", prob, fmt='0%', fill=input_fill)
    put(ws, f"F{r}", f"=(D{r}/B$3)^(1/3)-1", fmt=PCT)
    put(ws, f"G{r}", f"=F{r}+B$4", fmt=PCT)
    r += 1
put(ws, "A10", "Probability-weighted", font=bold)
put(ws, "D10", "=SUMPRODUCT(D7:D9,E7:E9)", fmt=PS, font=bold)
put(ws, "E10", "=SUM(E7:E9)", fmt='0%', font=bold)
put(ws, "F10", "=(D10/B$3)^(1/3)-1", fmt=PCT, font=bold)
put(ws, "G10", "=F10+B$4", fmt=PCT, font=bold)
for c in "ABCDEFG":
    ws[f"{c}10"].fill = total_fill

put(ws, "A12", "Probabilities default to the post-autopsy weights (45/42/13): management already broke "
               "one AI-inflection promise (FY25 guided +3-5%, came in −2.6%), so the Q3-FY26 catalyst gets "
               "a credibility haircut. Original weights were 40/45/15. Edit column E to test. "
               "NOTE: 3-yr FV here = present intrinsic value (the gap to fair value closes over 3 years, "
               "no further value compounding) — more conservative than model.md's forward targets that "
               "compound intrinsic value, so this shows ~+7% total vs the model's ~+9% at the same weights.",
    font=note_font)
ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A12:G14")

# ----------------------------------------------------------------------------
# VALUE LENSES
# ----------------------------------------------------------------------------
ws = wb.create_sheet("ValueLenses")
ws.sheet_view.showGridLines = False
for col, w in {"A": 38, "B": 18, "C": 18}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Graham / Buffett value lenses", font=title_font)

put(ws, "A3", "EARNINGS POWER VALUE (Greenwald/Graham — zero growth)", font=sub_font)
put(ws, "B4", "Charging recurring SBC"); put(ws, "C4", "Not charging SBC")
style_header_row(ws, 4, 2, start=2)
epv = [
    ("Adjusted EBITDA", 60.0, 60.0, MONEY, True),
    ("less: D&A", 5.2, 5.2, MONEY, True),
    ("less: recurring stock-comp", 10.9, 0.0, MONEY, True),
    ("Normalized EBIT", "=B5-B6-B7", "=C5-C6-C7", MONEY, False),
    ("Cash tax rate", 0.26, 0.26, PCT, True),
    ("NOPAT", "=B8*(1-B9)", "=C8*(1-C9)", MONEY, False),
    ("WACC", 0.105, 0.105, PCT, True),
    ("EPV operating (NOPAT/WACC)", "=B10/B11", "=C10/C11", MONEY, False),
    ("less: net debt", 72.7, 72.7, MONEY, True),
    ("EPV equity", "=B12-B13", "=C12-C13", MONEY, False),
    ("Shares (M)", 25.18, 25.18, '#,##0.00', True),
    ("EPV per share ($)", "=B14/B15", "=C14/C15", PS, False),
    ("vs. price $11.61", "=B16/11.61-1", "=C16/11.61-1", PCT, False),
]
r = 5
for label, b, c, fmt, is_input in epv:
    put(ws, f"A{r}", label, font=bold if "EPV per share" in label else None)
    put(ws, f"B{r}", b, fmt=fmt, fill=input_fill if is_input else None,
        font=bold if "EPV per share" in label else None)
    put(ws, f"C{r}", c, fmt=fmt, fill=input_fill if is_input else None,
        font=bold if "EPV per share" in label else None)
    if "EPV per share" in label:
        ws[f"A{r}"].fill = total_fill; ws[f"B{r}"].fill = total_fill; ws[f"C{r}"].fill = total_fill
    r += 1

put(ws, "A19", "REVERSE DCF (what the current price implies)", font=sub_font)
rev = [
    ("Current price ($)", 11.61, PS, True),
    ("Shares (M)", 25.18, '#,##0.00', True),
    ("Market cap ($M)", "=B20*B21", MONEY, False),
    ("Net debt ($M)", 72.7, MONEY, True),
    ("Enterprise value ($M)", "=B22+B23", MONEY, False),
    ("Normalized FCF ($M)", 32.4, MONEY, True),
    ("WACC", 0.105, PCT, True),
    ("Implied perpetual FCF growth", "=B26-B25/B24", PCT, False),
]
r = 20
for label, v, fmt, is_input in rev:
    put(ws, f"A{r}", label, font=bold if "Implied" in label else None)
    put(ws, f"B{r}", v, fmt=fmt, fill=input_fill if is_input else None,
        font=bold if "Implied" in label else None)
    if "Implied" in label:
        ws[f"A{r}"].fill = total_fill; ws[f"B{r}"].fill = total_fill
    r += 1
put(ws, "A28", "Nominal GDP ~4%. The market is pricing HCKT to grow at <half of GDP forever — "
               "the quantitative version of 'priced as a slowly-melting body shop.'", font=note_font)
ws["A28"].alignment = Alignment(wrap_text=True, vertical="top"); ws.merge_cells("A28:C29")

put(ws, "A31", "OWNER-EARNINGS YIELD (Buffett)", font=sub_font)
put(ws, "A32", "Normalized FCF / owner earnings ($M)"); put(ws, "B32", 32.4, fmt=MONEY, fill=input_fill)
put(ws, "A33", "Owner-earnings yield (on market cap)", font=bold)
put(ws, "B33", "=B32/B22", fmt=PCT, font=bold); ws["A33"].fill = total_fill; ws["B33"].fill = total_fill
put(ws, "A34", "FCF / enterprise value"); put(ws, "B34", "=B32/B24", fmt=PCT)

put(ws, "A36", "GRAHAM ASSET / NET-NET TEST", font=sub_font)
put(ws, "A37", "Net financial debt ($M)"); put(ws, "B37", 72.7, fmt=MONEY)
put(ws, "A38", "Goodwill + intangibles ($M)"); put(ws, "B38", 93.1, fmt=MONEY)
put(ws, "A39", "Asset floor"); put(ws, "B39", "NONE", font=red_font)
put(ws, "A40", "Net-debt, goodwill-heavy services business — no balance-sheet floor under the equity. "
               "The bear is a real earnings-multiple drawdown, not an asset-protected floor (the key "
               "contrast with USNA).", font=note_font)
ws["A40"].alignment = Alignment(wrap_text=True, vertical="top"); ws.merge_cells("A40:C41")

put(ws, "A43", "GRAHAM 'STRONG FINANCIAL CONDITION' TEST (Defensive Investor)", font=sub_font)
put(ws, "A44", "Current assets ($M)"); put(ws, "B44", 81.7, fmt=MONEY, fill=input_fill)
put(ws, "A45", "Current liabilities ($M)"); put(ws, "B45", 43.5, fmt=MONEY, fill=input_fill)
put(ws, "A46", "Current ratio  (Graham test: >= 2.0)", font=bold)
put(ws, "B46", "=B44/B45", fmt='0.00"x"', font=bold)
put(ws, "C46", '=IF(B46>=2,"PASS","FAIL")', font=red_font)
put(ws, "A47", "Working capital (net current assets, $M)")
put(ws, "B47", "=B44-B45", fmt=MONEY)
put(ws, "A48", "Long-term debt ($M)"); put(ws, "B48", 78.8, fmt=MONEY, fill=input_fill)
put(ws, "A49", "LT debt / working capital  (Graham test: <= 1.0)", font=bold)
put(ws, "B49", "=B48/B47", fmt='0.00"x"', font=bold)
put(ws, "C49", '=IF(B48<=B47,"PASS","FAIL")', font=red_font)
put(ws, "A50", "Source: Q1 FY26 10-Q balance sheet (XBRL AssetsCurrent $81.7M, LiabilitiesCurrent $43.5M, "
               "LongTermDebtNoncurrent $78.8M). BOTH prongs FAIL: current ratio 1.88x (< 2.0), and LT debt "
               "$78.8M > working capital $38.2M (2.1x). Strong cash flow, but a static balance-sheet safety "
               "screen HCKT does not pass — the downside is not asset-protected.", font=note_font)
ws["A50"].alignment = Alignment(wrap_text=True, vertical="top"); ws.merge_cells("A50:C52")

# ----------------------------------------------------------------------------
# SENSITIVITY
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Sensitivity")
ws.sheet_view.showGridLines = False
for col, w in {"A": 30, "B": 14, "C": 14, "D": 14}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Sensitivity tornado (base FV ≈ $14.50)", font=title_font)
put(ws, "A2", "Each driver flexed alone; values are model output from data/research/hckt/model.md §4.",
    font=note_font)
for j, h in enumerate(["Driver", "FV low ($)", "FV high ($)", "Swing ($/sh)"]):
    put(ws, f"{chr(65+j)}4", h)
style_header_row(ws, 4, 4)
tor = [
    ("Revenue trajectory (Q3 inflection?)", 8.6, 21.0),
    ("Exit / blended EBITDA multiple", 9.4, 19.8),
    ("Adjusted EBITDA margin", 11.2, 17.6),
    ("Stock-comp treatment", 11.8, 16.2),
    ("WACC", 12.6, 15.8),
    ("Cash tax rate", 13.4, 14.7),
]
r = 5
for name, lo, hi in tor:
    put(ws, f"A{r}", name)
    put(ws, f"B{r}", lo, fmt=PS)
    put(ws, f"C{r}", hi, fmt=PS)
    put(ws, f"D{r}", f"=C{r}-B{r}", fmt=PS)
    r += 1
ws.conditional_formatting.add("D5:D10",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=13,
                color=RED, showValue=True))
put(ws, "A12", "The answer hangs on the top two — both resolved by the same Q3-FY26 print: does revenue "
               "inflect, and does the market therefore re-rate off the 6x body-shop multiple toward the "
               "9-13x advisory band.", font=note_font)
ws["A12"].alignment = Alignment(wrap_text=True, vertical="top"); ws.merge_cells("A12:D13")

# ----------------------------------------------------------------------------
# PEERS  (comp set behind the SOTP/value multiples — read live from peer CSV)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Peers")
ws.sheet_view.showGridLines = False
for col, w in {"A": 9, "B": 22, "C": 11, "D": 11, "E": 10, "F": 11, "G": 12, "H": 11, "I": 60}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Peer comp set — multiples behind the SOTP & value lenses", font=title_font)
put(ws, "A2", "Source: data/research/hckt/peer_benchmarks.csv (stockanalysis.com + 10-K, as of 2026-05-29). "
              "Refresh against a live source before publication.", font=note_font)
heads = ["Ticker", "Name", "Inclusion", "EV/EBITDA", "EV/Rev", "Rev gr %",
         "Gross mgn", "EBITDA mgn", "Notes"]
for j, h in enumerate(heads):
    put(ws, f"{chr(65+j)}4", h)
style_header_row(ws, 4, len(heads))
peer_csv = Path(__file__).parent.parent / "data" / "research" / "hckt" / "peer_benchmarks.csv"
r = 5
with open(peer_csv, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        def num(k):
            v = row.get(k, "")
            try:
                return float(v)
            except (ValueError, TypeError):
                return None
        put(ws, f"A{r}", row["peer_ticker"], font=bold if row["inclusion"] == "subject" else None)
        put(ws, f"B{r}", row["peer_name"])
        put(ws, f"C{r}", row["inclusion"])
        if num("ev_ebitda") is not None:
            put(ws, f"D{r}", num("ev_ebitda"), fmt=MULT)
        if num("ev_revenue") is not None:
            put(ws, f"E{r}", num("ev_revenue"), fmt=MULT)
        if num("revenue_growth_pct") is not None:
            put(ws, f"F{r}", num("revenue_growth_pct") / 100, fmt=PCT)
        if num("gross_margin") is not None:
            put(ws, f"G{r}", num("gross_margin") / 100, fmt=PCT)
        if num("ebitda_margin") is not None:
            put(ws, f"H{r}", num("ebitda_margin") / 100, fmt=PCT)
        note = row.get("notes", "")
        put(ws, f"I{r}", note[:300], font=note_font)
        if row["inclusion"] == "subject":
            for c in "ABCDEFGH":
                ws[f"{c}{r}"].fill = total_fill
        r += 1
put(ws, f"A{r+1}", "Key: 'primary' = tight advisory comps (CRAI/FCN/HURN ~10.8-13.2x = the re-rate target); "
                   "'secondary' = directional; FORR (0.35x rev) = the AI-disintermediation bear comp; "
                   "EPAM (6x) = the body-shop floor the market currently applies. HCKT EV/EBITDA is 9.4x on "
                   "GAAP EBITDA but ~6.0x on adjusted EBITDA — the gap is the one-time stock-comp charge.",
    font=note_font)
ws[f"A{r+1}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{r+1}:I{r+2}")

# ----------------------------------------------------------------------------
# ASSUMPTIONS  (how each judgment input was constructed — WACC built live)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Assumptions")
ws.sheet_view.showGridLines = False
for col, w in {"A": 30, "B": 12, "C": 16, "D": 64}.items():
    ws.column_dimensions[col].width = w
put(ws, "A1", "Assumptions & judgment log — how the non-cited inputs were built", font=title_font)

put(ws, "A3", "WACC BUILD (CAPM + explicit judgment overlay)", font=sub_font)
wacc = [
    ("Risk-free rate (10-yr UST)", 0.043, PCT, "Market-observable, May 2026."),
    ("Beta (levered)", 0.976, '0.000', "yfinance — noisy on a thin-float micro-cap; don't defend the 3rd digit."),
    ("Equity risk premium", 0.055, PCT, "Convention (Damodaran ~4.5-5.5%). An assumption, not derived."),
    ("Size premium (micro-cap)", 0.015, PCT, "Duff & Phelps style, ~$292M cap. Judgment."),
    ("Cost of equity (CAPM)", "=B4+B5*B6+B7", PCT, "= rf + beta x ERP + size premium."),
    ("", None, None, ""),
    ("Pre-tax cost of debt", 0.06, PCT, "Revolver at ~SOFR + spread."),
    ("Tax rate", 0.26, PCT, "Normalized cash rate (see log below)."),
    ("After-tax cost of debt", "=B10*(1-B11)", PCT, "= pre-tax x (1 - tax)."),
    ("", None, None, ""),
    ("Market cap ($M)", 292.4, MONEY, "Equity market value."),
    ("Gross financial debt ($M)", 78.8, MONEY, "Revolver (market-value weight)."),
    ("Equity weight", "=B14/(B14+B15)", PCT, "= mkt cap / (mkt cap + debt)."),
    ("Debt weight", "=B15/(B14+B15)", PCT, ""),
    ("", None, None, ""),
    ("WACC — mechanical CAPM", "=B16*B8+B17*B12", PCT, "What the textbook build produces (~9.8%)."),
    ("Judgment overlay", 0.0075, PCT, "Illiquidity + thin coverage + single binary catalyst → raise the "
                                      "required return above the mechanical CAPM. RAISES the discount rate = "
                                      "lowers value = conservative direction."),
    ("WACC — applied (DCF base)", "=B19+B20", PCT, "The 10.5% used in the DCF base. Bull ~9.5% (≈ clean CAPM, "
                                                   "no overlay); bear 12% (heavier risk premium)."),
]
r = 4
for label, val, fmt, logic in wacc:
    if not label:
        r += 1
        continue
    is_calc = isinstance(val, str) and val.startswith("=")
    boldit = "applied" in label or "mechanical" in label
    put(ws, f"A{r}", label, font=bold if boldit else None)
    put(ws, f"B{r}", val, fmt=fmt, fill=None if is_calc else input_fill,
        font=bold if boldit else None)
    put(ws, f"D{r}", logic, font=note_font)
    if "applied" in label:
        for c in "AB":
            ws[f"{c}{r}"].fill = total_fill
    r += 1

put(ws, "A24", "JUDGMENT-INPUT LOG (every non-cited assumption: tier + how it was built)", font=sub_font)
for j, h in enumerate(["Input", "Base", "Tier", "How constructed / logic"]):
    put(ws, f"{chr(65+j)}25", h)
style_header_row(ws, 25, 4)
log = [
    ("WACC", "10.5%", "Built + judgment", "CAPM ~9.8% + 0.75pt illiquidity/binary-catalyst overlay (built above)."),
    ("Revenue growth path", "−4%→+4%", "Scenario judgment", "The binary 'does Q3-FY26 inflect' call. NOT a bottoms-up "
                                                             "bookings model — HCKT discloses no backlog/pipeline. The #1 swing ($12.4/sh)."),
    ("Adj. EBITDA margin", "20%", "Cited anchor", "FY25 actual 19.6% / Q1'26 20.3%; held flat — does NOT credit the +500bps AI lift."),
    ("Cash tax rate", "26%", "Normalized", "Mgmt guide 26.6% + FY23 actual 25.8%; deliberately NOT the optical 40.7% book rate (one-time SBC §162(m))."),
    ("Terminal growth", "+1%", "Judgment", "GDP-minus; ~45% (Oracle/SAP) commoditizing offsets IP growth. Cut from the agent's +2% in lead review."),
    ("Capex", "$6M", "Judgment", "Normalized from an elevated FY25 $7.9M toward the $4-6M history."),
    ("Working-capital use", "$1.5M", "Judgment", "Modest drag as revenue grows; asset-light services."),
    ("SOTP segment multiples", "9 / 6.5 / 6x", "Judgment", "Per-segment peer multiples (Peers tab): S&BT 9x (vs CRAI/HURN 10.8-13x), Oracle 6.5x, SAP 6x."),
    ("Scenario probabilities", "45/42/13", "Judgment", "Post-autopsy weights: management already broke one AI-inflection promise (FY25 guided +3-5%, came −2.6%) → heavier bear."),
]
r = 26
for inp, val, tier, logic in log:
    put(ws, f"A{r}", inp)
    put(ws, f"B{r}", val)
    put(ws, f"C{r}", tier)
    put(ws, f"D{r}", logic, font=note_font)
    ws[f"D{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

wb.save(OUT)
print("wrote", OUT)
