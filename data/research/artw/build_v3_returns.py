import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.load_workbook("data/research/artw/ARTW_model.xlsx")
for s in ["v3_Summary", "LBO_Returns", "LBO_Sensitivity"]:
    if s in wb.sheetnames:
        del wb[s]

B = Font(bold=True)
NAVY = Font(bold=True, color="FFFFFF")
fill = PatternFill("solid", fgColor="0A2540")
tint = PatternFill("solid", fgColor="FEF7F7")
yellow = PatternFill("solid", fgColor="FFF6BF")
MONEYX = '0.00"x"'

def title(ws, t):
    ws["A1"] = t
    ws["A1"].font = Font(bold=True, size=13, color="0A2540")

# ---------------- v3_Summary ----------------
ws = wb.create_sheet("v3_Summary", 0)
title(ws, "ARTW - v3 Control-Basis Recalibration + Search-Fund LBO Returns")
ws["A2"] = ("Companion to model.md v3 (2026-05-31). v2 sheets unchanged. "
            "Supporting: wacc_research.md, filing_review.md, acquisition_premium_comps.md")
ws["A2"].font = Font(italic=True, size=9)
rows = [
 ["v3 KEY RECALIBRATIONS (premise: friendly control / take-private)", "value", "note"],
 ["WACC base", "12.5%", "13.5%->12.5%; strip minority-illiquidity slice for a control buyer. bear 14% / bull 11.5%"],
 ["Standalone corp-cost add-back", "$0.30M", "$0.45M->$0.30M; take-private removes public-co cost. Standalone EBITDA $1.55M->$1.70M"],
 ["Near-term cash tax", "~0% yrs1-3", "NOL ~$7.1M gross fed (no VA); 382 caps post-change -> modest holiday only"],
 ["Ag disposal (base)", "$7.7M", "UNCHANGED (reviewer-disciplined); new floor evidence argues conservative"],
 ["", "", ""],
 ["RECALIBRATED INTRINSIC (control basis)", "FV basic", "FV diluted", "vs $2.58"],
 ["Bear", "$1.10", "$1.00", "-57%"],
 ["Base", "$2.80", "$2.55", "+8% / -1%"],
 ["Bull", "$5.19", "$4.74", "+101%"],
 ["Prob-weighted (30/45/25)", "$2.89", "$2.63", "+12% / +2%"],
 ["diluted = 5.684M sh (+500K equity plan overhang)", "", "", ""],
 ["", "", ""],
 ["SEARCH-FUND LBO RETURNS (recommended ~$10M / 46% leverage, 5-yr)", "MOIC", "IRR", ""],
 ["Bear", "0.39x", "-17%", "bounded loss; debt collateral-covered, no wipeout"],
 ["Base", "1.73x", "+12%", "de-levered modular FCF + ag-funded delever"],
 ["Bull", "3.73x", "+30%", "modular grows + re-rates to ~10x"],
 ["", "", ""],
 ["VERDICT", ("Asymmetric, asset-protected search acquisition. Base earns a real ~12% IRR; upside scales "
             "with the modular re-rate; bear is a bounded, debt-covered loss. Cap leverage ~$10M "
             "(>55% wipes the bear). Conviction hinges on the modular leg-split (data floor).")],
]
for i, r in enumerate(rows, start=4):
    for j, v in enumerate(r):
        ws.cell(i, j + 1, v)
for cell in ["A4", "A11", "A18", "A23"]:
    ws[cell].font = B
for col, w in {"A": 44, "B": 12, "C": 14, "D": 50}.items():
    ws.column_dimensions[col].width = w

# ---------------- LBO_Returns (formula-driven) ----------------
ws = wb.create_sheet("LBO_Returns")
title(ws, "Search-Fund LBO Returns - formula-driven (yellow = inputs)")
assum = [
 (3, "ASSUMPTIONS", None), (4, "Current price ($/sh)", 2.58), (5, "Entry premium", 0.10),
 (6, "Entry price ($/sh)", "=B4*(1+B5)"), (7, "Shares (M)", 5.184),
 (8, "Equity purchase ($M)", "=B6*B7"), (9, "Refinance existing net debt ($M)", 6.40),
 (10, "Transaction fees ($M)", "=0.03*B8"), (11, "TOTAL USES ($M)", "=B8+B9+B10"),
 (13, "Acquisition debt - ABL+SBA+seller ($M)", 10.0),
 (14, "Sponsor equity ($M)", "=B11-B13"), (15, "Blended debt rate", 0.0625),
]
for r, lab, val in assum:
    ws.cell(r, 1, lab)
    if val is not None:
        ws.cell(r, 2, val)
for cell in ["A3", "A11", "B11", "A14"]:
    ws[cell].font = B
for cell in ["B4", "B5", "B13"]:
    ws[cell].fill = yellow
ws["B5"].number_format = "0%"
ws["B15"].number_format = "0.00%"

def block(start, name, growths, margin, exitm, ag):
    r0 = start
    ws.cell(r0, 1, "%s CASE - operating & debt schedule  (margin %.1f%%, exit %.0fx, ag $%sM)"
            % (name, margin * 100, exitm, ag)).font = B
    ry = r0 + 1
    ws.cell(ry, 1, "Year")
    for k, yr in enumerate(["Entry(0)", 1, 2, 3, 4, 5]):
        ws.cell(ry, 3 + k, yr).font = B
    L = ["g", "rev", "eb", "ebit", "tax", "cap", "wc", "fcf", "db", "int", "ag", "av", "pd", "de", "cb", "ce"]
    labels = {"g": "Revenue growth", "rev": "Revenue ($M)", "eb": "EBITDA ($M)", "ebit": "EBIT ($M)",
              "tax": "Cash tax ($M)", "cap": "Capex ($M)", "wc": "Chg Working capital ($M)", "fcf": "Unlevered FCF ($M)",
              "db": "Debt - beginning ($M)", "int": "Interest ($M)", "ag": "Ag disposal proceeds ($M)",
              "av": "Cash available ($M)", "pd": "Debt paydown ($M)", "de": "Debt - ending ($M)",
              "cb": "Cash - beginning ($M)", "ce": "Cash - ending ($M)"}
    rr = {k: r0 + 2 + i for i, k in enumerate(L)}
    for k in L:
        ws.cell(rr[k], 1, labels[k])
    cols = ["D", "E", "F", "G", "H"]
    for i, c in enumerate(cols):
        cell = ws["%s%d" % (c, rr["g"])]
        cell.value = growths[i]
        cell.number_format = "0%"
        cell.fill = yellow
    ws["C%d" % rr["rev"]] = 10.226
    prev = "C"
    for c in cols:
        ws["%s%d" % (c, rr["rev"])] = "=%s%d*(1+%s%d)" % (prev, rr["rev"], c, rr["g"])
        ws["%s%d" % (c, rr["eb"])] = "=%s%d*%s" % (c, rr["rev"], margin)
        ws["%s%d" % (c, rr["ebit"])] = "=%s%d-0.025*%s%d" % (c, rr["eb"], c, rr["rev"])
        ws["%s%d" % (c, rr["cap"])] = "=0.025*%s%d" % (c, rr["rev"])
        ws["%s%d" % (c, rr["wc"])] = "=0.10*(%s%d-%s%d)" % (c, rr["rev"], prev, rr["rev"])
        ws["%s%d" % (c, rr["fcf"])] = "=%s%d-%s%d-%s%d-%s%d" % (c, rr["eb"], c, rr["tax"], c, rr["cap"], c, rr["wc"])
        ws["%s%d" % (c, rr["int"])] = "=%s%d*$B$15" % (c, rr["db"])
        ws["%s%d" % (c, rr["av"])] = "=%s%d-%s%d+%s%d" % (c, rr["fcf"], c, rr["int"], c, rr["ag"])
        ws["%s%d" % (c, rr["pd"])] = "=MIN(%s%d,MAX(%s%d,0))" % (c, rr["db"], c, rr["av"])
        ws["%s%d" % (c, rr["de"])] = "=%s%d-%s%d" % (c, rr["db"], c, rr["pd"])
        ws["%s%d" % (c, rr["ce"])] = "=%s%d+(%s%d-%s%d)" % (c, rr["cb"], c, rr["av"], c, rr["pd"])
        prev = c
    for c in ["D", "E", "F"]:
        ws["%s%d" % (c, rr["tax"])] = 0
    for c in ["G", "H"]:
        ws["%s%d" % (c, rr["tax"])] = "=0.21*%s%d" % (c, rr["ebit"])
    ws["D%d" % rr["db"]] = "=$B$13"
    for p, c in zip(["D", "E", "F", "G"], ["E", "F", "G", "H"]):
        ws["%s%d" % (c, rr["db"])] = "=%s%d" % (p, rr["de"])
    ws["D%d" % rr["ag"]] = ag
    for c in ["E", "F", "G", "H"]:
        ws["%s%d" % (c, rr["ag"])] = 0
    ws["D%d" % rr["cb"]] = 0
    for p, c in zip(["D", "E", "F", "G"], ["E", "F", "G", "H"]):
        ws["%s%d" % (c, rr["cb"])] = "=%s%d" % (p, rr["ce"])
    re = r0 + 20
    ws.cell(re, 1, "Exit EV ($M) = Yr5 EBITDA x multiple").font = B
    ws.cell(re, 2, "=H%d*%s" % (rr["eb"], exitm))
    ws.cell(re + 1, 1, "  less ending debt / plus ending cash")
    ws.cell(re + 2, 1, "Exit equity ($M)").font = B
    ws.cell(re + 2, 2, "=B%d-H%d+H%d" % (re, rr["de"], rr["ce"]))
    ws.cell(re + 3, 1, "MOIC").font = B
    ws.cell(re + 3, 2, "=B%d/$B$14" % (re + 2))
    ws["B%d" % (re + 3)].number_format = MONEYX
    ws.cell(re + 4, 1, "IRR (5-yr)").font = B
    ws.cell(re + 4, 2, '=IF(B%d>0,(B%d/$B$14)^(1/5)-1,"neg")' % (re + 2, re + 2))
    ws["B%d" % (re + 4)].number_format = "0%"
    return re + 4

e = block(18, "BASE", [.08, .06, .05, .04, .04], 0.166, 7.0, 7.7)
e = block(e + 3, "BULL", [.18, .14, .10, .08, .06], 0.19, 10.0, 10.4)
e = block(e + 3, "BEAR", [-.10, -.05, 0, .02, .03], 0.13, 5.0, 5.5)
for col, w in {"A": 34, "B": 12, "C": 11, "D": 9, "E": 9, "F": 9, "G": 9, "H": 9}.items():
    ws.column_dimensions[col].width = w

# ---------------- LBO_Sensitivity ----------------
ws = wb.create_sheet("LBO_Sensitivity")
title(ws, "LBO Sensitivity & Downside Coverage")
ws["A3"] = "LEVERAGE SENSITIVITY (MOIC / IRR; 5-yr hold, entry 10% premium)"
ws["A3"].font = B
hd = ["Acq debt ($M)", "Sponsor eq ($M)", "Bear MOIC", "Bear IRR", "Base MOIC", "Base IRR", "Bull MOIC", "Bull IRR"]
for j, h in enumerate(hd):
    c = ws.cell(4, j + 1, h)
    c.font = NAVY
    c.fill = fill
data = [
 [14.5, 7.0, "-0.2x", "wipe", "2.01x", "15%", "5.40x", "40%"],
 [12.0, 9.6, "0.19x", "-28%", "1.84x", "13%", "4.29x", "34%"],
 [10.0, 11.6, "0.39x", "-17%", "1.73x", "12%", "3.73x", "30%"],
 [8.0, 13.6, "0.53x", "-12%", "1.64x", "10%", "3.34x", "27%"],
]
for i, row in enumerate(data, start=5):
    for j, v in enumerate(row):
        c = ws.cell(i, j + 1, v)
        if abs(row[0] - 10.0) < 0.01:
            c.fill = tint
            c.font = B
ws.cell(9, 1, "Recommended ~$10M (<=50%): bear survives (debt collateral-covered); over $12M wipes the bear.").font = Font(italic=True, size=9)

ws["A11"] = "DOWNSIDE COVERAGE AT CLOSE (bear-liquidation collateral vs entry debt)"
ws["A11"].font = B
cov = [["Asset", "Carrying $M", "Bear recovery %", "Liquidation $M"],
       ["Ag inventory", 9.5, 0.50, "=B13*C13"], ["Accounts receivable", 1.6, 0.85, "=B14*C14"],
       ["PP&E (plants/land)", 4.5, 0.35, "=B15*C15"], ["Modular (going-concern floor)", 7.0, 1.0, "=B16*C16"],
       ["TOTAL COLLATERAL", "", "", "=SUM(D13:D16)"],
       ["Entry debt ($M)", "", "", 10.0], ["Coverage (x)", "", "", "=D17/D18"]]
for i, row in enumerate(cov, start=12):
    for j, v in enumerate(row):
        c = ws.cell(i, j + 1, v)
        if i == 12:
            c.font = NAVY
            c.fill = fill
for cell in ["C13", "C14", "C15", "C16"]:
    ws[cell].number_format = "0%"
ws["D19"].number_format = MONEYX
ws["A17"].font = B
ws["A19"].font = B
for col, w in {"A": 40, "B": 13, "C": 15, "D": 14}.items():
    ws.column_dimensions[col].width = w

wb.save("data/research/artw/ARTW_model.xlsx")
print("Saved. Sheets:", wb.sheetnames)
